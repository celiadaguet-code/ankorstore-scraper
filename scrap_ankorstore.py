#!/usr/bin/env python3
"""
scrap_ankorstore.py
====================

Pipeline complet :
    URL marque WooCommerce  ->  .xlsx Ankorstore prêt à importer

Étapes internes :
    1. Detection : Store API v1 (avec fallback Store API legacy)
    2. Récupère TOUS les produits (paginé, par 100)
    3. Récupère les variations en batch (?type=variation&include=...)
    4. Filtre ateliers / consultations / cartes cadeaux
    5. Nettoie les descriptions (Divi + HTML + entités)
    6. Classifie les attributs (taille / couleur / autre)
    7. Génère SKU stables quand absents ({slug}-{variation_id}, max 50 chars)
    8. Écrit le .xlsx au format Ankorstore (1 ligne = 1 variante)
    9. Produit un rapport .log par marque + .csv agrégé

Usage :
    python3 scrap_ankorstore.py https://cosmella.fr
    python3 scrap_ankorstore.py https://cosmella.fr --max 20
    python3 scrap_ankorstore.py https://marque1.fr https://marque2.fr

Setup auto :
    Au premier lancement, si openpyxl n'est pas dispo, le script crée un venv
    local (./venv/) et s'y relance tout seul. Aucune action manuelle requise.
"""

from __future__ import annotations

# =============================================================================
# SECTION 1 — Bootstrap venv (ne pas modifier)
# =============================================================================
# Cette section s'exécute avant tout autre import non-stdlib. Elle s'assure
# qu'openpyxl est disponible, sinon elle crée un venv ./venv/ à côté du script,
# installe les deps, et relance le script avec le Python du venv.
import os
import sys


def _bootstrap_venv() -> None:
    """Crée un venv local si openpyxl absent, puis re-exec le script dedans."""
    try:
        import openpyxl  # noqa: F401
        return  # Déjà OK
    except ImportError:
        pass

    script_dir = os.path.dirname(os.path.abspath(__file__))
    venv_dir = os.path.join(script_dir, "venv")
    venv_python = os.path.join(venv_dir, "bin", "python3")

    # Si on est déjà dans le venv mais sans openpyxl, c'est un autre problème
    if os.path.realpath(sys.executable).startswith(os.path.realpath(venv_dir)):
        print("[FATAL] openpyxl absent même dans le venv, install probablement échouée.",
              file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(venv_python):
        print("[setup] Création du venv local ./venv/ ...", flush=True)
        import venv as _venv
        _venv.EnvBuilder(with_pip=True, clear=False).create(venv_dir)

    print("[setup] Installation d'openpyxl dans le venv ...", flush=True)
    import subprocess
    subprocess.check_call([venv_python, "-m", "pip", "install", "--quiet", "openpyxl"])

    print("[setup] Relance du script dans le venv ...\n", flush=True)
    # Re-exécute en remplaçant le process courant
    os.execv(venv_python, [venv_python, os.path.abspath(__file__), *sys.argv[1:]])


_bootstrap_venv()


# =============================================================================
# SECTION 2 — Imports (après bootstrap)
# =============================================================================
import argparse
import csv
import gzip
import html
import json
import logging
import re
import shutil
import ssl
import time
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# SECTION 3 — Configuration
# =============================================================================

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
HTTP_TIMEOUT = 20
CRAWL_DELAY = 1.2  # secondes entre 2 requêtes vers le même domaine
HTTP_RETRIES = 2
VARIATION_BATCH = 25  # variations fetchées par appel ?include=...
SKU_MAX_LEN = 50

# Mots-clés titre qui font filtrer un produit (ateliers, gift cards...)
NON_MERCHANDISE_KEYWORDS_TITLE = (
    "atelier", "workshop", "consultation", "rendez-vous", "rdv",
    "gift card", "chèque cadeau", "bon cadeau", "carte cadeau",
    "séance", "stage", "cours", "formation", "billet", "ticket",
)

# Mots-clés catégorie qui font filtrer
NON_MERCHANDISE_KEYWORDS_CATEGORY = (
    "atelier", "workshop", "service", "consultation",
    "carte cadeau", "gift", "événement", "evenement",
)

# Classification des noms d'attributs (insensible casse/accents)
SIZE_KEYWORDS = ("taille", "size", "sizes", "tailles", "pointure", "größe", "talla", "dimension")
COLOR_KEYWORDS = ("couleur", "color", "colour", "couleurs", "coloris", "teinte", "farbe", "colore")

# Mapping pays texte → ISO-2 (couvre les cas courants ; le template a un onglet
# "Codes pays" exhaustif pour les cas exotiques)
COUNTRY_TO_ISO = {
    "france": "FR", "français": "FR", "française": "FR", "fabriqué en france": "FR",
    "italie": "IT", "italy": "IT", "italia": "IT",
    "espagne": "ES", "spain": "ES", "españa": "ES",
    "portugal": "PT",
    "allemagne": "DE", "germany": "DE", "deutschland": "DE",
    "royaume-uni": "GB", "uk": "GB", "united kingdom": "GB", "angleterre": "GB",
    "belgique": "BE", "belgium": "BE",
    "pays-bas": "NL", "netherlands": "NL", "hollande": "NL",
    "suisse": "CH", "switzerland": "CH",
    "chine": "CN", "china": "CN",
    "inde": "IN", "india": "IN",
    "maroc": "MA", "morocco": "MA",
    "tunisie": "TN", "tunisia": "TN",
    "états-unis": "US", "etats-unis": "US", "usa": "US", "united states": "US",
    "japon": "JP", "japan": "JP",
}

# Détection des tags booléens à partir du nom + description + catégories
# Format : (col, anywhere_keywords, name_or_category_only_keywords, negatives)
# - anywhere : mots-clés cherchés dans nom + description + catégories + tags
# - name_or_category_only : mots-clés cherchés UNIQUEMENT dans nom + catégories
#   (pour éviter les faux positifs sur des sous-chaînes communes type
#   "emporter" matché par "porter")
# - negatives : si présents, désactivent le tag même si positives matchent
BOOLEAN_TAGS = [
    ("Meilleure vente",
        ("meilleure vente", "best seller", "bestseller", "top vente"),
        (),
        ()),
    ("Contient de l'alcool",
        (
            # Mention directe — sans ambiguïté, cherchable partout
            "alcool", "alcohol", "alcoolisé", "alcoolique", "alcoholic",
            # Beer brewing termes peu ambigus
            "brasserie", "brassée", "brassez", "brassé", "brewery",
            # Vins effervescents très spécifiques
            "champagne", "crémant", "cremant", "prosecco", "mousseux",
            # Spiritueux non-ambigus
            "spiritueux", "cognac", "armagnac", "calvados", "kirsch",
            "mezcal", "absinthe", "pastis", "anisette", "eau-de-vie",
            "schnaps",
        ),
        (
            # Mots ambigus — uniquement dans NOM ou CATÉGORIE pour éviter
            # les faux positifs dans le corps de description (ex: "emporter"
            # contient "porter", "saison" = "season", "vin" dans "vintage")
            "bière", "biere", "beer", "vin", "vins", "wine", "vino",
            "ipa", "neipa", "stout", "porter", "ale", "lager",
            "pilsner", "pils", "kolsch", "kölsch", "saison",
            "rhum", "rum", "whisky", "whiskey", "scotch",
            "gin", "vodka", "tequila", "liqueur",
            "cidre", "cider", "poiré", "cava",
        ),
        (
            # Faux positifs explicites
            "sans alcool", "no alcohol", "alcohol free", "alcohol-free",
            "non-alcoolisé", "non alcoolisé", "0% alcool", "0 % alcool",
            "non-alcoholic", "free of alcohol",
        )),
    ("Sans cruauté",
        ("cruelty free", "cruelty-free", "sans cruauté", "non testé sur animaux"),
        (),
        ()),
    ("Écologique",
        ("écologique", "ecologique", "eco-friendly", "ecofriendly", "écoresponsable",
         "ecoresponsable", "durable", "sustainable"),
        (),
        ()),
    ("Fait main",
        ("fait main", "fait à la main", "handmade", "hand made", "hand-made", "artisanal"),
        (),
        ()),
    ("Biologique",
        # "organique" RETIRÉ : trop ambigu en français (sens chimique : "pierre
        # précieuse organique", "matière organique", "chimie organique" — sans
        # rapport avec l'agriculture/cosmétique bio).
        ("biologique", "organic", "bio", "label bio"),
        (),
        ()),
    ("Végan",
        ("vegan", "végan", "vegane", "végane"),
        (),
        ()),
    ("Objectif zéro déchet",
        ("zero waste", "zéro déchet", "zero déchet", "0 déchet"),
        (),
        ()),
]

# Colonnes du template Ankorstore (ordre exact, 38 colonnes)
ANKORSTORE_COLUMNS = [
    "SKU", "Nom du produit", "Description du produit",
    "Tailles des variantes", "Couleurs des variants", "Autres attributs de variante",
    "Image de la variante",
    "Image 1", "Image 2", "Image 3", "Image 4", "Image 5",
    "Prix de gros/unité", "Prix de détail/unité", "Taux de TVA %",
    "Remise sur le prix de gros %", "Nombre d'unités par paquet", "Stock",
    "Fabriqué en", "Code douanier", "IAN (EAN-13)",
    "Dimensions", "Volume", "Poids",
    "Composition", "Liste INCI", "Matériau",
    "Durée de vie garantie à réception", "Température de conservation",
    "Liste des ingrédients",
    "Meilleure vente", "Contient de l'alcool", "Sans cruauté",
    "Écologique", "Fait main", "Biologique", "Végan", "Objectif zéro déchet",
]


# =============================================================================
# SECTION 4 — Logging
# =============================================================================

def setup_logger(brand_slug: str, output_dir: Path) -> logging.Logger:
    logger = logging.getLogger(f"scrap.{brand_slug}")
    logger.handlers.clear()
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter("%(asctime)s %(levelname)-7s %(message)s", "%H:%M:%S")

    # Console : INFO+
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # File : DEBUG (tout)
    log_path = output_dir / f"{brand_slug}.log"
    fh = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    logger.debug(f"Log file: {log_path}")
    return logger


# =============================================================================
# SECTION 5 — Couche HTTP (stdlib)
# =============================================================================

def _build_ssl_context(verify: bool):
    if verify:
        return ssl.create_default_context()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


_SSL_VERIFY: bool | None = None


def _ssl_works() -> bool:
    """Détecte une fois si la vérification SSL native est dispo (Python homebrew = non)."""
    global _SSL_VERIFY
    if _SSL_VERIFY is not None:
        return _SSL_VERIFY
    try:
        req = urllib.request.Request("https://www.google.com/", headers={"User-Agent": USER_AGENT})
        urllib.request.urlopen(req, timeout=5, context=_build_ssl_context(True)).close()
        _SSL_VERIFY = True
    except Exception:
        _SSL_VERIFY = False
    return _SSL_VERIFY


class HTTPError(Exception):
    pass


def http_get_json(url: str, logger: logging.Logger, retries: int = HTTP_RETRIES) -> tuple[int, dict, Any]:
    """GET avec retries exponentiels. Retourne (status, headers, parsed_json_or_text_or_None).

    Le header Accept est volontairement permissif : HTML/XML/JSON tous bienvenus.
    Le parsing JSON automatique se déclenche UNIQUEMENT si le Content-Type de la
    réponse contient 'json' (laissé tel quel sinon).
    """
    # NOTE : ne PAS inclure "application/json" dans Accept. Sur certains
    # serveurs PrestaShop (cas lagazellemarrakchia.com), si Accept contient
    # application/json, le serveur retourne 200 OK avec un body vide. Bizarre
    # mais reproduit en local. On laisse le serveur décider via Content-Type.
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, identity",
    }
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers, method="GET")
            ctx = _build_ssl_context(verify=_ssl_works())
            with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT, context=ctx) as resp:
                raw = resp.read()
                if resp.headers.get("Content-Encoding", "").lower() == "gzip":
                    try:
                        raw = gzip.decompress(raw)
                    except Exception:
                        pass
                hdrs = {k: v for k, v in resp.headers.items()}
                ct = resp.headers.get("Content-Type", "").lower()
                text = raw.decode("utf-8", errors="replace") if raw else ""
                if "json" in ct:
                    try:
                        return resp.status, hdrs, json.loads(text)
                    except json.JSONDecodeError as e:
                        logger.debug(f"JSON decode error on {url}: {e}")
                        return resp.status, hdrs, text
                return resp.status, hdrs, text
        except urllib.error.HTTPError as e:
            # 4xx : pas de retry
            if 400 <= e.code < 500:
                try:
                    body = e.read().decode("utf-8", errors="replace")[:300]
                except Exception:
                    body = ""
                logger.debug(f"HTTP {e.code} on {url}: {body!r}")
                return e.code, dict(e.headers or {}), None
            last_err = e
            logger.debug(f"HTTP {e.code} on {url}, retry...")
        except Exception as e:
            last_err = e
            logger.debug(f"network error on {url}: {type(e).__name__}: {e}, retry...")
        if attempt < retries - 1:
            time.sleep(1.5 ** attempt)
    raise HTTPError(f"Échec après {retries} tentatives sur {url} : {last_err}")


# =============================================================================
# SECTION 6 — Scraper WooCommerce
# =============================================================================

@dataclass
class WooProduct:
    """Représentation brute d'un produit Woo (ce qu'on a récupéré)."""
    id: int
    name: str
    slug: str
    type: str  # 'simple' | 'variable' | 'grouped' | ...
    sku: str
    short_description: str
    description: str
    permalink: str
    prices: dict
    images: list[dict]
    categories: list[dict]
    tags: list[dict]
    attributes: list[dict]  # noms + terms possibles
    variations_ids: list[int]  # IDs des variations
    variations_data: list[dict]  # détails fetchés des variations (id, prices, sku, attributes)
    is_in_stock: bool
    low_stock_remaining: int | None
    raw: dict  # tout le JSON brut, au cas où
    features: list[dict] = field(default_factory=list)  # [{"name": "Matière", "value": "Coton"}, ...]


class WooScraper:
    """Scrape un catalogue WooCommerce avec cascade Tier 1 (Store API) → Tier 2 (sitemap+HTML)."""

    def __init__(self, brand_url: str, logger: logging.Logger):
        self.brand_url = brand_url
        p = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
        self.base = f"{p.scheme}://{p.netloc}"
        self.domain = p.netloc.replace("www.", "")
        self.logger = logger
        self.api_root = f"{self.base}/wp-json/wc/store/v1"
        self._brand_suffix = None  # détecté en Tier 2 pour strip du titre

    # -- Détection du tier qui marche ------------------------------------

    def _try_api(self, root: str) -> tuple[bool, int]:
        """Teste un endpoint Store API. Retourne (works, total)."""
        url = f"{root}/products?per_page=1"
        try:
            status, hdrs, data = http_get_json(url, self.logger)
        except HTTPError as e:
            self.logger.debug(f"API {root} indisponible : {e}")
            return False, 0
        if status == 200 and isinstance(data, list):
            total = int(hdrs.get("X-WP-Total") or hdrs.get("x-wp-total") or len(data))
            return True, total
        self.logger.debug(f"API {root} a renvoyé status={status} (pas une liste JSON)")
        return False, 0

    def select_tier(self) -> str:
        """Retourne 'api_v1', 'api_legacy', ou 'sitemap'."""
        ok, total = self._try_api(self.api_root)
        if ok:
            self.logger.info(f"Tier 1 — Store API v1 OK ({total} produits déclarés)")
            return "api_v1"

        legacy = f"{self.base}/wp-json/wc/store"
        ok, total = self._try_api(legacy)
        if ok:
            self.logger.warning(f"Tier 1bis — fallback Store API legacy ({total} produits)")
            self.api_root = legacy
            return "api_legacy"

        self.logger.warning("Tier 1 et 1bis indisponibles — bascule en Tier 2 (sitemap+HTML)")
        return "sitemap"

    # -- Liste paginée ----------------------------------------------------

    def fetch_all_products(self, max_products: int | None = None) -> list[dict]:
        """Pagine /products en récupérant per_page=100 jusqu'à total."""
        per_page = 100
        page = 1
        all_products: list[dict] = []
        while True:
            url = f"{self.api_root}/products?per_page={per_page}&page={page}"
            self.logger.info(f"GET liste page {page} (per_page={per_page})")
            status, hdrs, data = http_get_json(url, self.logger)
            if status != 200 or not isinstance(data, list):
                self.logger.error(f"page {page} échouée (status={status})")
                break
            all_products.extend(data)
            self.logger.info(f"  -> +{len(data)} produits (cumulé : {len(all_products)})")
            if len(data) < per_page:
                break
            page += 1
            time.sleep(CRAWL_DELAY)
            if max_products and len(all_products) >= max_products:
                all_products = all_products[:max_products]
                break
        return all_products

    # -- Variations en batch ---------------------------------------------

    def fetch_variations(self, variation_ids: list[int]) -> dict[int, dict]:
        """Fetch toutes les variations via ?type=variation&include=...
        Retourne un dict {variation_id: variation_data}.
        """
        out: dict[int, dict] = {}
        for i in range(0, len(variation_ids), VARIATION_BATCH):
            batch = variation_ids[i:i + VARIATION_BATCH]
            url = (f"{self.api_root}/products"
                   f"?type=variation&include={','.join(map(str, batch))}"
                   f"&per_page={len(batch)}")
            self.logger.info(f"GET variations batch {i // VARIATION_BATCH + 1} "
                             f"({len(batch)} IDs)")
            status, hdrs, data = http_get_json(url, self.logger)
            if isinstance(data, list):
                for obj in data:
                    if isinstance(obj, dict) and "id" in obj:
                        out[obj["id"]] = obj
                self.logger.debug(f"  -> {len(data)}/{len(batch)} variations récupérées")
            else:
                self.logger.warning(f"  -> batch échoué (status={status})")
            time.sleep(CRAWL_DELAY)
        return out

    # -- Construction des WooProduct -------------------------------------

    def build(self, max_products: int | None = None) -> list[WooProduct]:
        tier = self.select_tier()
        if tier == "sitemap":
            return self._build_via_sitemap(max_products=max_products)
        # Sinon Tier 1 / 1bis = même code (api_root déjà ajusté)
        return self._build_via_api(max_products=max_products)

    # =================================================================
    # Tier 2 — Sitemap + HTML scraping (JSON-LD + data-product_variations)
    # =================================================================

    def _find_product_sitemap(self) -> list[str]:
        """Trouve les sitemaps produits (peut y en avoir plusieurs pour gros catalogues).

        Stratégie : essaie /sitemap_index.xml, puis /sitemap.xml, puis /wp-sitemap.xml.
        Cherche les <loc> contenant 'product' dans leur chemin.
        """
        candidate_indexes = [
            f"{self.base}/sitemap_index.xml",
            f"{self.base}/sitemap.xml",
            f"{self.base}/wp-sitemap.xml",
        ]
        for url in candidate_indexes:
            try:
                status, hdrs, body = http_get_json(url, self.logger)
            except HTTPError:
                continue
            if status != 200 or not isinstance(body, str):
                continue
            self.logger.debug(f"Sitemap index trouvé : {url}")
            locs = re.findall(r"<loc>([^<]+)</loc>", body)
            product_sms = [u for u in locs if "product" in u.lower()]
            if product_sms:
                self.logger.info(f"{len(product_sms)} product-sitemap(s) détecté(s) "
                                 f"via {url}")
                return product_sms
            # Pas de sub-sitemap product : peut-être que le sitemap actuel
            # contient déjà directement les URLs produits ?
            if any("/product/" in u or "/produit/" in u or "/boutique/" in u for u in locs):
                self.logger.info(f"Sitemap {url} contient directement des URLs produits")
                return [url]
        return []

    def _collect_product_urls(self, sitemap_urls: list[str]) -> list[str]:
        """Pour chaque sitemap, extrait les URLs produits."""
        product_urls: list[str] = []
        for sm in sitemap_urls:
            try:
                status, hdrs, body = http_get_json(sm, self.logger)
            except HTTPError as e:
                self.logger.warning(f"Sitemap {sm} indisponible : {e}")
                continue
            if status != 200 or not isinstance(body, str):
                continue
            locs = re.findall(r"<loc>([^<]+)</loc>", body)
            # Filtre : on retire les URLs qui ne sont clairement PAS des produits
            EXCLUDE_PATTERNS = (
                "/categorie-produit/", "/product-category/", "/category/", "/categories/",
                "/tag/", "/etiquette/", "/page/", "/author/", "/auteur/",
                "/boutique/?", "/shop/?",
            )
            locs = [u for u in locs if not any(seg in u for seg in EXCLUDE_PATTERNS)]
            # On veut idéalement des URLs qui ressemblent à des produits ; si on a
            # un motif clair on filtre, sinon on prend tout (sitemap déjà spécifique)
            INCLUDE_PATTERNS = ("/product/", "/produit/", "/products/", "/produits/")
            filtered = [u for u in locs if any(seg in u for seg in INCLUDE_PATTERNS)]
            if not filtered:
                # Sitemap déjà spécifique aux produits (cas WooCommerce native sitemap)
                # On exclut juste les URLs qui finissent par "/" sans slug détaillé
                # (page d'index "/boutique/" par exemple)
                filtered = [u for u in locs if u.rstrip("/").count("/") > 3]
            product_urls.extend(filtered)
            self.logger.debug(f"{sm} -> {len(filtered)} URLs produits")
            time.sleep(0.5)
        # Dédoublonne en gardant l'ordre
        product_urls = list(dict.fromkeys(product_urls))
        return product_urls

    def _fetch_html(self, url: str) -> str:
        """Fetch une page HTML (réutilise http_get_json qui retourne le texte si non-JSON)."""
        status, hdrs, body = http_get_json(url, self.logger)
        if status != 200 or not isinstance(body, str):
            raise HTTPError(f"Page {url} status={status}")
        return body

    # ---- Helpers d'extraction du HTML --------------------------------

    @staticmethod
    def _extract_jsonld_product(html_text: str) -> dict | None:
        """Parcourt tous les blocs JSON-LD, retourne le premier de type Product."""
        for block in re.findall(
            r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html_text, flags=re.DOTALL | re.IGNORECASE,
        ):
            try:
                data = json.loads(block.strip())
            except json.JSONDecodeError:
                continue
            candidates = []
            if isinstance(data, list):
                candidates = data
            elif isinstance(data, dict):
                if "@graph" in data and isinstance(data["@graph"], list):
                    candidates = data["@graph"]
                else:
                    candidates = [data]
            for c in candidates:
                if isinstance(c, dict):
                    t = c.get("@type")
                    if t == "Product" or (isinstance(t, list) and "Product" in t):
                        return c
        return None

    @staticmethod
    def _extract_variations_data(html_text: str) -> list[dict] | None:
        """Extrait data-product_variations (WooCommerce variations JSON)."""
        for p in (
            r'data-product_variations\s*=\s*"([^"]+)"',
            r"data-product_variations\s*=\s*'([^']+)'",
            r'data-product_variations\s*=\s*&quot;(.+?)&quot;',
        ):
            m = re.search(p, html_text, flags=re.DOTALL)
            if m:
                try:
                    return json.loads(html.unescape(m.group(1)))
                except json.JSONDecodeError:
                    return None
        return None

    def _detect_brand_suffix(self, html_text: str) -> str:
        """Détecte le suffixe site dans le titre (ex: ' • Le comptoir des saveurs').
        Méthode : regarde le og:site_name ou le <title> de la home.
        Cached après 1er appel.
        """
        if self._brand_suffix is not None:
            return self._brand_suffix
        # Cherche og:site_name dans la page
        m = re.search(
            r'<meta[^>]+property=["\']og:site_name["\'][^>]+content=["\']([^"\']+)["\']',
            html_text,
        )
        site_name = m.group(1).strip() if m else ""
        self._brand_suffix = site_name
        if site_name:
            self.logger.debug(f"og:site_name détecté : {site_name!r}")
        return site_name

    def _clean_jsonld_name(self, raw_name: str, html_text: str) -> str:
        """Nettoie le nom : décode HTML, strip le suffixe site (• Le comptoir...)."""
        n = html.unescape(raw_name or "").strip()
        site = self._detect_brand_suffix(html_text)
        if site:
            # Patterns courants : " • Site", " | Site", " - Site", " – Site"
            for sep in (" • ", " | ", " - ", " – "):
                if n.endswith(f"{sep}{site}"):
                    n = n[: -(len(sep) + len(site))].strip()
                    break
            # Cas où le séparateur est attaché au site_name lui-même
            if n.endswith(site):
                n = n[: -len(site)].rstrip(" •|-–").strip()
        return n

    @staticmethod
    def _jsonld_images(image_field: Any) -> list[dict]:
        """Normalise le champ image du JSON-LD vers une liste de {src, alt}."""
        if not image_field:
            return []
        items = image_field if isinstance(image_field, list) else [image_field]
        out = []
        for it in items:
            if isinstance(it, str):
                out.append({"src": it})
            elif isinstance(it, dict):
                url = it.get("url") or it.get("contentUrl") or it.get("src")
                if url:
                    out.append({"src": url, "alt": it.get("caption") or it.get("name") or ""})
        return out

    @staticmethod
    def _extract_full_description_from_page(html_text: str) -> str:
        """Extrait la description longue depuis le HTML de la page produit.

        WooCommerce expose la description longue dans plusieurs endroits selon
        le thème. On essaie en cascade.
        """
        patterns = [
            # WooCommerce standard : onglet "Description"
            r'<div[^>]*\bid=["\']tab-description["\'][^>]*>(.*?)</div>\s*(?:<div[^>]*\bid=["\']tab-|<section|</section|<footer|<div[^>]*woocommerce-tabs__panel-end)',
            r'<div[^>]*\bclass="[^"]*woocommerce-Tabs-panel--description[^"]*"[^>]*>(.*?)</div>\s*(?:<div[^>]*woocommerce-Tabs-panel|<section|</section|<footer)',
            # itemprop="description"
            r'<div[^>]*\bitemprop="description"[^>]*>(.*?)</div>',
            # class="product-description"
            r'<div[^>]*\bclass="[^"]*product-description[^"]*"[^>]*>(.*?)</div>',
        ]
        for p in patterns:
            m = re.search(p, html_text, flags=re.DOTALL | re.IGNORECASE)
            if m:
                return m.group(1)
        return ""

    @staticmethod
    def _humanize_attribute_key(key: str) -> str:
        """`attribute_pa_color` ou `attribute_couleur_principale` -> 'Couleur Principale'."""
        k = key
        for prefix in ("attribute_pa_", "attribute_"):
            if k.startswith(prefix):
                k = k[len(prefix):]
                break
        k = k.replace("_", " ").replace("-", " ").strip()
        return k.title()

    def _extract_product_from_html(self, url: str, html_text: str) -> WooProduct | None:
        """Construit un WooProduct à partir d'une page HTML produit."""
        ld = self._extract_jsonld_product(html_text)
        variations_data = self._extract_variations_data(html_text)

        if not ld and not variations_data:
            self.logger.warning(f"[Tier2] {url} : ni JSON-LD ni variations data trouvés")
            return None

        # ID stable depuis l'URL si JSON-LD n'en fournit pas
        slug_match = re.search(r"/(?:produit|product|shop|boutique)/([^/?#]+)", url)
        slug = slug_match.group(1) if slug_match else url.rstrip("/").rsplit("/", 1)[-1]
        product_id = abs(hash(slug)) % (10 ** 9)  # ID synthétique stable
        if ld and "@id" in ld:
            # Si JSON-LD a un @id, on tente d'extraire un id numérique
            m_id = re.search(r"product/(\d+)", ld.get("@id", ""))
            if m_id:
                product_id = int(m_id.group(1))

        name = ""
        description = ""
        sku = ""
        images: list[dict] = []
        categories: list[dict] = []
        offers_price = None
        offers_currency = "EUR"
        in_stock = True

        if ld:
            name = self._clean_jsonld_name(ld.get("name") or "", html_text)
            description = html.unescape(ld.get("description") or "").strip()
            sku = (ld.get("sku") or "").strip()
            images = self._jsonld_images(ld.get("image"))
            cat = ld.get("category")
            if isinstance(cat, str):
                categories = [{"name": html.unescape(cat).strip()}]
            elif isinstance(cat, list):
                categories = [{"name": html.unescape(c).strip()} for c in cat if isinstance(c, str)]
            offers = ld.get("offers")
            if isinstance(offers, dict):
                # Offer simple
                if offers.get("@type") == "Offer":
                    offers_price = offers.get("price")
                    offers_currency = offers.get("priceCurrency") or "EUR"
                    avail = offers.get("availability", "")
                    in_stock = "InStock" in avail
                # AggregateOffer (variable product) — on prend lowPrice par défaut
                elif offers.get("@type") == "AggregateOffer":
                    offers_price = offers.get("lowPrice")
                    offers_currency = offers.get("priceCurrency") or "EUR"
                    avail = offers.get("availability", "")
                    in_stock = "InStock" in avail

        # Fallback HTML pour titre si JSON-LD vide
        if not name:
            m = re.search(
                r'<h1[^>]*class="[^"]*product_title[^"]*"[^>]*>(.*?)</h1>',
                html_text, flags=re.DOTALL,
            )
            if m:
                name = re.sub(r"<[^>]+>", "", html.unescape(m.group(1))).strip()

        # Construit les variations_data au format identique à ce que renvoie l'API
        # On reformat data-product_variations -> structure compatible
        variations_data_compat: list[dict] = []
        variations_ids: list[int] = []
        if variations_data:
            for v in variations_data:
                vid = v.get("variation_id")
                try:
                    vid = int(vid) if vid is not None else None
                except (ValueError, TypeError):
                    vid = None
                if vid is None:
                    continue
                variations_ids.append(vid)
                v_sku = (v.get("sku") or "").strip()
                v_price = v.get("display_price")
                v_regular = v.get("display_regular_price")
                # Currency_minor_unit conventionnel : 2
                # display_price est en unité majeure (€), pas en centimes !
                # On convertit en centimes pour rester cohérent avec le reste du pipeline.
                def _to_cents(x: Any) -> str | None:
                    if x in (None, "", False):
                        return None
                    try:
                        return str(int(round(float(x) * 100)))
                    except (ValueError, TypeError):
                        return None
                v_price_cents = _to_cents(v_price)
                v_regular_cents = _to_cents(v_regular)
                # Stock
                v_in_stock_raw = v.get("is_in_stock")
                if isinstance(v_in_stock_raw, str):
                    v_in_stock = v_in_stock_raw.lower() in ("true", "1", "yes")
                else:
                    v_in_stock = bool(v_in_stock_raw)
                # Image variation
                v_image = v.get("image") or {}
                v_imgs = []
                if isinstance(v_image, dict):
                    src = v_image.get("url") or v_image.get("full_src") or v_image.get("src")
                    if src:
                        v_imgs.append({"src": src, "alt": v_image.get("alt", "")})
                # Attributs : {"attribute_xxx": "value"} -> [{"name": "Xxx", "value": "value"}]
                v_attrs = []
                for k, val in (v.get("attributes") or {}).items():
                    if not val:
                        continue
                    human = self._humanize_attribute_key(k)
                    v_attrs.append({"name": human, "value": val})
                variations_data_compat.append({
                    "id": vid,
                    "sku": v_sku,
                    "prices": {
                        "price": v_price_cents,
                        "regular_price": v_regular_cents or v_price_cents,
                        "currency_code": offers_currency or "EUR",
                        "currency_minor_unit": 2,
                    },
                    "is_in_stock": v_in_stock,
                    "low_stock_remaining": None,
                    "images": v_imgs,
                    "_parent_attributes": v_attrs,
                })

        # Stock parent : si variable on regarde si au moins 1 variation in_stock
        if variations_data_compat:
            in_stock = any(v["is_in_stock"] for v in variations_data_compat)

        # Prix parent en centimes (pour produits simples)
        parent_price_cents = None
        if offers_price is not None:
            try:
                parent_price_cents = str(int(round(float(offers_price) * 100)))
            except (ValueError, TypeError):
                pass

        # Détermine type : variable si on a des variations, sinon simple
        p_type = "variable" if variations_data_compat else "simple"

        # Description longue depuis le HTML de la page (le JSON-LD ne contient
        # généralement que la description courte)
        full_description_html = self._extract_full_description_from_page(html_text)

        return WooProduct(
            id=product_id,
            name=name,
            slug=slug,
            type=p_type,
            sku=sku,
            short_description=description,  # JSON-LD description (courte)
            description=full_description_html,  # HTML brut, sera nettoyé en aval
            permalink=url,
            prices={
                "price": parent_price_cents,
                "currency_code": offers_currency,
                "currency_minor_unit": 2,
            } if parent_price_cents else {},
            images=images,
            categories=categories,
            tags=[],
            attributes=[],
            variations_ids=variations_ids,
            variations_data=variations_data_compat,
            is_in_stock=in_stock,
            low_stock_remaining=None,
            raw={"_tier": 2, "_url": url},
        )

    def _build_via_sitemap(self, max_products: int | None = None) -> list[WooProduct]:
        """Tier 2 : sitemap → HTML pages → WooProducts."""
        sitemap_urls = self._find_product_sitemap()
        if not sitemap_urls:
            raise HTTPError(f"Tier 2 : aucun sitemap produits trouvé sur {self.base}")

        product_urls = self._collect_product_urls(sitemap_urls)
        self.logger.info(f"Tier 2 : {len(product_urls)} URLs produits à scraper")
        if max_products:
            product_urls = product_urls[:max_products]
            self.logger.info(f"  (limité à {len(product_urls)} pour ce run)")

        products: list[WooProduct] = []
        failures = 0
        for i, url in enumerate(product_urls, 1):
            if i % 10 == 0 or i == len(product_urls):
                self.logger.info(f"  page {i}/{len(product_urls)} ({failures} échecs)")
            try:
                html_text = self._fetch_html(url)
                product = self._extract_product_from_html(url, html_text)
                if product:
                    products.append(product)
                else:
                    failures += 1
            except Exception as e:
                failures += 1
                self.logger.warning(f"[Tier2] {url} échec : {type(e).__name__}: {e}")
            time.sleep(CRAWL_DELAY)
        if failures:
            self.logger.warning(f"Tier 2 terminé avec {failures}/{len(product_urls)} échecs")
        return products

    # =================================================================
    # Tier 1 — Construction des WooProduct depuis la Store API
    # =================================================================

    def _build_via_api(self, max_products: int | None = None) -> list[WooProduct]:
        raw_products = self.fetch_all_products(max_products=max_products)

        # Collecte les IDs de variations à fetcher
        all_var_ids: list[int] = []
        for p in raw_products:
            if p.get("type") == "variable":
                for v in (p.get("variations") or []):
                    vid = v.get("id") if isinstance(v, dict) else v
                    if isinstance(vid, int):
                        all_var_ids.append(vid)
        all_var_ids = list(dict.fromkeys(all_var_ids))
        self.logger.info(f"{len(all_var_ids)} variations uniques à fetcher")

        variations_map = self.fetch_variations(all_var_ids) if all_var_ids else {}

        # Assemble les WooProduct
        products: list[WooProduct] = []
        for p in raw_products:
            var_ids = []
            var_data = []
            for v in (p.get("variations") or []):
                if isinstance(v, dict):
                    vid = v.get("id")
                    var_ids.append(vid)
                    detail = variations_map.get(vid)
                    if detail:
                        # On ajoute l'attribut "named" depuis le parent
                        # (la variation seule ne contient pas les noms d'attribut)
                        merged = dict(detail)
                        merged["_parent_attributes"] = v.get("attributes") or []
                        var_data.append(merged)
                else:
                    var_ids.append(v)
            products.append(WooProduct(
                id=p.get("id"),
                name=p.get("name") or "",
                slug=p.get("slug") or "",
                type=p.get("type") or "simple",
                sku=p.get("sku") or "",
                short_description=p.get("short_description") or "",
                description=p.get("description") or "",
                permalink=p.get("permalink") or "",
                prices=p.get("prices") or {},
                images=p.get("images") or [],
                categories=p.get("categories") or [],
                tags=p.get("tags") or [],
                attributes=p.get("attributes") or [],
                variations_ids=var_ids,
                variations_data=var_data,
                is_in_stock=bool(p.get("is_in_stock", True)),
                low_stock_remaining=p.get("low_stock_remaining"),
                raw=p,
            ))
        return products


# =============================================================================
# SECTION 6bis — Scraper PrestaShop (sitemap + HTML / data-product JSON)
# =============================================================================

class PrestaShopScraper:
    """Scrape un catalogue PrestaShop via sitemap + parsing HTML des pages produit.

    Stratégie :
      - Liste des produits via sitemap (avec plusieurs fallbacks)
      - Pour chaque produit : extraction depuis
          (a) attribut HTML `data-product="{...}"` (le plus riche, 95+ clés)
          (b) JSON-LD Product schema (complément, surtout pour les images)
          (c) sélecteurs HTML du form (pour les attributs/options de variantes)
      - Génère le produit cartésien des combinaisons d'attributs (1 ligne par combo)
    """

    def __init__(self, brand_url: str, logger: logging.Logger):
        self.brand_url = brand_url
        p = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
        self.base = f"{p.scheme}://{p.netloc}"
        self.domain = p.netloc.replace("www.", "")
        self.logger = logger
        # Détection langue depuis l'URL (ex: /fr/, /en/) — fallback ""
        path_parts = p.path.strip("/").split("/")
        self.lang_prefix = (
            f"/{path_parts[0]}" if path_parts and len(path_parts[0]) == 2 else ""
        )

    # -- Sitemap ---------------------------------------------------------

    def _find_sitemap(self) -> list[str]:
        """Cherche le sitemap. Retourne la liste des URLs trouvées dedans (produits)."""
        candidates = [
            f"{self.base}/sitemap.xml",
            f"{self.base}{self.lang_prefix}/sitemap.xml" if self.lang_prefix else None,
            f"{self.base}/1_fr_0_sitemap.xml",
            f"{self.base}/1_en_0_sitemap.xml",
            f"{self.base}/sitemap_index.xml",
            f"{self.base}/index.php?fc=module&module=gsitemap&controller=sitemap",
        ]
        # Ajoute aussi les URLs annoncées par robots.txt
        try:
            status, hdrs, body = http_get_json(f"{self.base}/robots.txt", self.logger)
            if status == 200 and isinstance(body, str):
                for line in body.splitlines():
                    if line.lower().startswith("sitemap:"):
                        sm_url = line.split(":", 1)[1].strip()
                        if sm_url and sm_url not in candidates:
                            candidates.insert(0, sm_url)
        except Exception:
            pass

        candidates = [c for c in candidates if c]
        all_urls: list[str] = []
        for cand in candidates:
            try:
                status, hdrs, body = http_get_json(cand, self.logger)
            except HTTPError:
                continue
            if status != 200 or not isinstance(body, str):
                continue
            if not body.lstrip().startswith(("<?xml", "<urlset", "<sitemapindex")):
                continue
            self.logger.debug(f"Sitemap accessible : {cand}")
            # Sitemap index → suivre les sub-sitemaps
            if "<sitemapindex" in body[:300]:
                sub_locs = re.findall(r"<loc>([^<]+)</loc>", body)
                self.logger.info(f"Sitemap index avec {len(sub_locs)} sub-sitemaps")
                for sub in sub_locs:
                    # On garde tous les sub-sitemaps, on filtrera ensuite
                    try:
                        s2, _, body2 = http_get_json(sub, self.logger)
                    except HTTPError:
                        continue
                    if s2 == 200 and isinstance(body2, str):
                        all_urls.extend(re.findall(r"<loc>([^<]+)</loc>", body2))
                    time.sleep(0.3)
                if all_urls:
                    self.logger.info(f"Sitemap utilisé : {cand} (via index)")
                    return all_urls
            else:
                # Sitemap direct
                locs = re.findall(r"<loc>([^<]+)</loc>", body)
                if locs:
                    self.logger.info(f"Sitemap utilisé : {cand} ({len(locs)} URLs)")
                    return locs
        return []

    @staticmethod
    def _is_product_url(url: str) -> bool:
        """Heuristique : URL Presta de fiche produit."""
        # /{cat}/{id}-{slug}.html
        if re.search(r"/\d+(?:-\d+)?-[^/]+\.html?$", url):
            return True
        # /{cat}/{id}_{slug} (variant PS plus rare)
        return False

    @staticmethod
    def _is_category_url(url: str) -> bool:
        """L'URL ressemble à une catégorie Presta (/{id}-{slug} sans .html)."""
        parsed = urlparse(url)
        path = parsed.path.rstrip("/")
        if not path:
            return False  # racine = home, pas catégorie
        # Pattern catégorie : /{id_numerique}-{slug_letters}
        # On exclut explicitement les URLs produits (qui finissent en .html)
        if path.endswith(".html") or path.endswith(".htm"):
            return False
        return bool(re.search(r"/\d+(?:-\d+)?-[a-z0-9][a-z0-9\-_]+$", path, re.IGNORECASE))

    def _get_product_urls(self) -> list[str]:
        # MODE 1 — URL catégorie : on scope sur cette catégorie uniquement
        if self._is_category_url(self.brand_url):
            self.logger.info(
                f"Mode catégorie : scraping scopé à {self.brand_url} "
                f"(et ses pages paginées)"
            )
            return self._crawl_single_category(self.brand_url)

        # MODE 2 — URL home : on utilise le sitemap global de la marque
        urls = self._find_sitemap()
        kept = [u for u in urls if self.domain in u and self._is_product_url(u)] if urls else []
        if kept:
            return list(dict.fromkeys(kept))

        # MODE 3 — fallback : sitemap absent/vide → crawl home + catégories
        self.logger.warning(
            f"Aucune URL produit trouvée dans le sitemap. "
            f"Bascule en crawl alternatif (home + catégories)."
        )
        return self._crawl_for_product_urls()

    def _crawl_single_category(self, category_url: str, max_pages: int = 50) -> list[str]:
        """Crawl une page catégorie (et ses pages de pagination) → liste produits.

        Supporte la pagination PrestaShop : `?p=2`, `?p=3`, etc.
        Stoppe quand une page ne ramène plus de nouveaux produits.
        """
        product_urls: list[str] = []  # ordered list (pas set) pour reproductibilité
        seen: set[str] = set()
        page = 1

        while page <= max_pages:
            if page == 1:
                page_url = category_url
            else:
                sep = "&" if "?" in category_url else "?"
                page_url = f"{category_url}{sep}p={page}"

            self.logger.info(f"  catégorie page {page}: {page_url}")
            try:
                status, hdrs, body = http_get_json(page_url, self.logger)
            except HTTPError as e:
                self.logger.warning(f"Catégorie page {page} échec : {e}")
                break
            if status != 200 or not isinstance(body, str):
                self.logger.warning(f"Catégorie page {page} status={status}")
                break

            # Extrait les URLs produits du HTML
            page_added = 0
            for href in re.findall(r'href=["\']([^"\']+)["\']', body):
                # Normalise en URL absolue
                if href.startswith("//"):
                    href = f"https:{href}"
                elif href.startswith("/"):
                    href = f"{self.base}{href}"
                elif not href.startswith("http"):
                    continue
                if self.domain not in href:
                    continue
                # Strip fragments / query (sauf si c'est la pagination)
                href_clean = href.split("#")[0].split("?")[0]
                if self._is_product_url(href_clean) and href_clean not in seen:
                    seen.add(href_clean)
                    product_urls.append(href_clean)
                    page_added += 1

            self.logger.info(f"    +{page_added} produits (total: {len(product_urls)})")
            if page_added == 0:
                # Plus de nouveaux produits → fin de pagination
                break
            page += 1
            time.sleep(CRAWL_DELAY)

        self.logger.info(f"Crawl catégorie terminé : {len(product_urls)} produits uniques")
        return product_urls

    # ---- Fallback : crawl home + catégories ---------------------------

    def _crawl_for_product_urls(self, max_categories: int = 20) -> list[str]:
        """Fallback quand le sitemap est vide : crawl la home et les pages
        catégories pour trouver les URLs produits.

        Stratégie :
          1. Fetch home -> extrait URLs produits directes + URLs catégories
          2. Si on a déjà >= 20 produits, retourne
          3. Sinon, fetch chaque URL catégorie et extrait les produits
        """
        product_urls: set[str] = set()
        category_urls: set[str] = set()

        # 1. Home page
        try:
            status, hdrs, body = http_get_json(self.brand_url, self.logger)
        except HTTPError as e:
            self.logger.error(f"Home inaccessible : {e}")
            return []
        if status != 200 or not isinstance(body, str):
            return []

        # Extrait toutes les URLs du body
        all_links = re.findall(r'href=["\']([^"\']+)["\']', body)
        for href in all_links:
            # Normalise en URL absolue
            if href.startswith("//"):
                href = f"https:{href}"
            elif href.startswith("/"):
                href = f"{self.base}{href}"
            elif not href.startswith("http"):
                continue
            if self.domain not in href:
                continue
            # Strip fragments / query strings pour dédoublonnage
            href_clean = href.split("#")[0].split("?")[0]
            if self._is_product_url(href_clean):
                product_urls.add(href_clean)
            # URL catégorie typique : /{n}-{slug} ou /{slug} sans .html
            elif (
                re.search(r"/\d+-[a-z0-9\-_]+/?$", href_clean)
                and not href_clean.endswith(".html")
                and "/category/" not in href_clean
                and "?" not in href_clean
            ):
                category_urls.add(href_clean)

        self.logger.info(
            f"Crawl home : {len(product_urls)} produits directs, "
            f"{len(category_urls)} catégories candidates"
        )

        # 2. Si on a déjà assez, retourne
        if len(product_urls) >= 20:
            self.logger.info(f"Assez de produits trouvés directement ({len(product_urls)})")
            return list(product_urls)

        # 3. Sinon, crawl les catégories
        cats_to_visit = list(category_urls)[:max_categories]
        for i, cat_url in enumerate(cats_to_visit, 1):
            self.logger.info(f"  catégorie {i}/{len(cats_to_visit)} : {cat_url}")
            try:
                status, hdrs, body = http_get_json(cat_url, self.logger)
            except HTTPError:
                continue
            if status != 200 or not isinstance(body, str):
                continue
            for href in re.findall(r'href=["\']([^"\']+)["\']', body):
                if href.startswith("//"):
                    href = f"https:{href}"
                elif href.startswith("/"):
                    href = f"{self.base}{href}"
                elif not href.startswith("http"):
                    continue
                if self.domain not in href:
                    continue
                href_clean = href.split("#")[0].split("?")[0]
                if self._is_product_url(href_clean):
                    product_urls.add(href_clean)
            time.sleep(CRAWL_DELAY)

        self.logger.info(f"Crawl alternatif terminé : {len(product_urls)} URLs produits")
        return list(product_urls)

    # -- Extraction d'une page produit ----------------------------------

    @staticmethod
    def _extract_data_product(html_text: str) -> dict | None:
        """Récupère le contenu de l'attribut HTML `data-product="{...}"` (JSON encodé).

        Note : data-product est présent uniquement sur PrestaShop 1.7+.
        Pour les anciennes versions, voir _extract_legacy_presta_product().
        """
        m = re.search(r'data-product\s*=\s*(["\'])(\{.+?\})\1', html_text, flags=re.DOTALL)
        if not m:
            return None
        raw = m.group(2)
        # Décode les entités HTML (souvent &quot; au lieu de ")
        raw = html.unescape(raw)
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return None

    @staticmethod
    def _extract_legacy_presta_product(url: str, html_text: str) -> dict | None:
        """Fallback pour les vieux PrestaShop (1.5/1.6) sans `data-product`.

        Parse les sélecteurs HTML standards et retourne un dict avec la MÊME
        structure que `data-product` aurait retourné, pour rester compatible
        avec le pipeline d'extraction principal.
        """
        # Détection : on doit être sur une page produit (body class contient 'product-{N}')
        m = re.search(r'<body[^>]+class="[^"]*\bproduct-(\d+)\b', html_text)
        if not m:
            # Fallback : var product_page_product_id = 12345;
            m = re.search(r'product_page_product_id\s*=\s*(\d+)', html_text)
        if not m:
            return None
        try:
            product_id = int(m.group(1))
        except (ValueError, TypeError):
            return None

        # Titre depuis h1 (premier match)
        name = ""
        m = re.search(
            r'<h1[^>]*(?:class="[^"]*(?:pb-center-column|title|product-name)[^"]*"|itemprop="name")[^>]*>(.*?)</h1>',
            html_text, flags=re.DOTALL,
        )
        if not m:
            m = re.search(r'<h1[^>]*>(.*?)</h1>', html_text, flags=re.DOTALL)
        if m:
            name = re.sub(r"<[^>]+>", " ", html.unescape(m.group(1)))
            name = re.sub(r"\s+", " ", name).strip()

        # Prix depuis meta itemprop="price"
        price_amount = None
        m = re.search(r'<meta[^>]+itemprop="price"[^>]+content="([^"]+)"', html_text)
        if m:
            try:
                price_amount = float(m.group(1))
            except ValueError:
                pass
        if price_amount is None:
            # Fallback : data-product-price
            m = re.search(r'data-product-price="([^"]+)"', html_text)
            if m:
                try:
                    price_amount = float(m.group(1))
                except ValueError:
                    pass

        # Description courte
        description_short = ""
        for pat in (
            r'<div[^>]+id="short_description_content"[^>]*>(.*?)</div>',
            r'<div[^>]+id="short_description_block"[^>]*>(.*?)</div>',
            r'<section[^>]+class="[^"]*short-description[^"]*"[^>]*>(.*?)</section>',
        ):
            m = re.search(pat, html_text, flags=re.DOTALL)
            if m:
                description_short = m.group(1)
                break

        # Description longue (onglet description)
        description_long = ""
        for pat in (
            r'<section[^>]+id="description"[^>]*>(.*?)</section>',
            r'<div[^>]+id="idTab1"[^>]*>(.*?)(?=<div[^>]+id="idTab|<section\s|</body)',
            r'<div[^>]+itemprop="description"[^>]*>(.*?)</div>',
        ):
            m = re.search(pat, html_text, flags=re.DOTALL)
            if m:
                description_long = m.group(1)
                break

        # SKU / Référence
        sku = ""
        for pat in (
            r'<span[^>]+itemprop="sku"[^>]*>([^<]+)</span>',
            r'<div[^>]+id="product_reference"[^>]*>(?:[^<]|<(?!span))*?<span[^>]*>([^<]+)</span>',
            r'<span[^>]+id="product_reference"[^>]*>([^<]+)</span>',
        ):
            m = re.search(pat, html_text, flags=re.DOTALL)
            if m:
                sku = m.group(1).strip()
                break

        # Stock : généralement pas exposé dans le HTML legacy ; on assume "available"
        availability = "available"
        m = re.search(r'<link[^>]+itemprop="availability"[^>]+href="([^"]+)"', html_text)
        if m:
            avail_url = m.group(1)
            if "OutOfStock" in avail_url:
                availability = "out_of_stock"

        # Slug depuis URL
        slug_match = re.search(r"/\d+(?:-\d+)?-([^./]+)\.html?$", url)
        slug = slug_match.group(1) if slug_match else ""

        # Features depuis <table class="table-data-sheet">
        # Pattern : <tr><td>Nom</td><td>Valeur</td></tr>
        features: list[dict] = []
        m_table = re.search(
            r'<table[^>]*class="[^"]*table-data-sheet[^"]*"[^>]*>(.*?)</table>',
            html_text, flags=re.DOTALL | re.IGNORECASE,
        )
        if m_table:
            for row in re.findall(r'<tr[^>]*>(.*?)</tr>', m_table.group(1),
                                  flags=re.DOTALL):
                cells = re.findall(r'<td[^>]*>(.*?)</td>', row, flags=re.DOTALL)
                if len(cells) >= 2:
                    feat_name = re.sub(r"<[^>]+>", "", html.unescape(cells[0]))
                    feat_name = re.sub(r"\s+", " ", feat_name).strip()
                    feat_value = re.sub(r"<[^>]+>", "", html.unescape(cells[1]))
                    feat_value = re.sub(r"\s+", " ", feat_value).strip()
                    if feat_name and feat_value and feat_value.lower() not in ("sans", "non", "-", ""):
                        features.append({"name": feat_name, "value": feat_value})

        return {
            "id_product": product_id,
            "reference": sku,
            "name": name,
            "link_rewrite": slug,
            "description_short": description_short,
            "description": description_long or description_short,
            "price_amount": price_amount,
            "quantity": None,
            "availability": availability,
            "features": features,
            "attributes": {},  # legacy : pas de variantes structurées
        }

    @staticmethod
    def _extract_jsonld_product(html_text: str) -> dict | None:
        """Parcourt les blocs JSON-LD, retourne le premier de type Product."""
        for block in re.findall(
            r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html_text, flags=re.DOTALL | re.IGNORECASE,
        ):
            try:
                data = json.loads(block.strip())
            except json.JSONDecodeError:
                continue
            candidates = []
            if isinstance(data, list):
                candidates = data
            elif isinstance(data, dict):
                if "@graph" in data and isinstance(data["@graph"], list):
                    candidates = data["@graph"]
                else:
                    candidates = [data]
            for c in candidates:
                if isinstance(c, dict):
                    t = c.get("@type")
                    if t == "Product" or (isinstance(t, list) and "Product" in t):
                        return c
        return None

    @staticmethod
    def _extract_attribute_groups(
        html_text: str, dp: dict | None = None,
    ) -> list[tuple[str, list[tuple[str, str]]]]:
        """Récupère les groupes d'attributs depuis le form add-to-cart.

        Source primaire pour les noms de groupes : data-product.attributes[].group
        (PrestaShop expose le nom humain "Couleur", "Dimensions"...). Fallback :
        regex sur les labels HTML du form.

        Retourne : [ (nom_groupe, [(value_id, label_value), ...]), ... ]
        """
        # Isole le form
        form_match = re.search(
            r'<form[^>]+id=["\']add-to-cart-or-refresh["\'][^>]*>(.*?)</form>',
            html_text, flags=re.DOTALL | re.IGNORECASE,
        )
        if not form_match:
            return []
        form_html = form_match.group(1)

        groups: list[tuple[str, list[tuple[str, str]]]] = []

        # SOURCE PRIMAIRE : data-product.attributes[N].id_attribute_group -> .group
        # data-product expose un dict {numero_attr_actuel: {id_attribute_group, group, ...}}
        # qui nous donne le mapping {group_id: human_name}
        group_id_to_name: dict[str, str] = {}
        if dp and isinstance(dp.get("attributes"), dict):
            for _, attr_data in dp["attributes"].items():
                if isinstance(attr_data, dict):
                    grp_id = attr_data.get("id_attribute_group")
                    grp_name = attr_data.get("group")
                    if grp_id and grp_name:
                        group_id_to_name[str(grp_id)] = str(grp_name).strip().rstrip(":").strip().capitalize()

        # SOURCE SECONDAIRE : regex sur le HTML du form
        # On essaie plusieurs structures de label :
        #   - <label class="form-control-label|control-label">Couleur</label>
        #   - <span class="control-label">Couleur</span>
        #   - <h2 class="h6|product-variants-item__title">Couleur</h2>
        # Puis le select|input du groupe
        if not group_id_to_name:
            block_pattern = re.compile(
                r'(?:'
                r'<label[^>]*class="[^"]*(?:form-control-label|control-label)[^"]*"[^>]*>(.*?)</label>'
                r'|'
                r'<span[^>]*class="[^"]*control-label[^"]*"[^>]*>(.*?)</span>'
                r'|'
                r'<h[2-6][^>]*class="[^"]*(?:h6|product-variants-item__title)[^"]*"[^>]*>(.*?)</h[2-6]>'
                r')'
                r'(.*?)'
                r'(?:'
                r'<select[^>]+name=["\']group\[(\d+)\]["\']'
                r'|'
                r'<input[^>]+name=["\']group\[(\d+)\]["\']'
                r')',
                re.DOTALL,
            )
            for match in block_pattern.finditer(form_html):
                label_html = match.group(1) or match.group(2) or match.group(3) or ""
                label_clean = re.sub(r"<[^>]+>", "", html.unescape(label_html)).strip().rstrip(":").strip()
                gid = match.group(5) or match.group(6)
                if gid and label_clean:
                    group_id_to_name[gid] = label_clean.capitalize()

        # Extract <select> groups
        for sm in re.finditer(
            r'<select[^>]+name=["\']group\[(\d+)\]["\'][^>]*>(.*?)</select>',
            form_html, flags=re.DOTALL,
        ):
            gid = sm.group(1)
            sbody = sm.group(2)
            options = []
            for om in re.finditer(
                r'<option[^>]*value=["\']([^"\']+)["\'][^>]*>(.*?)</option>',
                sbody, flags=re.DOTALL,
            ):
                val = om.group(1)
                label = re.sub(r"<[^>]+>", "", html.unescape(om.group(2))).strip()
                if val and label and val != "0":
                    options.append((val, label))
            if options:
                gname = group_id_to_name.get(gid, f"Group {gid}")
                groups.append((gname, options))

        # Extract radio/checkbox groups (couleurs souvent en radio)
        radio_groups: dict[str, list[tuple[str, str]]] = {}
        for rm in re.finditer(
            r'<input[^>]+type=["\'](?:radio|checkbox)["\'][^>]+name=["\']group\[(\d+)\]["\'][^>]+value=["\']([^"\']+)["\'][^>]*(?:title=["\']([^"\']+)["\']|data-original-title=["\']([^"\']+)["\']|aria-label=["\']([^"\']+)["\'])?',
            form_html,
        ):
            gid = rm.group(1)
            val = rm.group(2)
            label = rm.group(3) or rm.group(4) or rm.group(5) or ""
            if not label:
                # Cherche le span associé après le input
                # Pattern : <input ...> ... <span ...>label</span>
                ctx_start = rm.end()
                ctx = form_html[ctx_start:ctx_start + 500]
                lm = re.search(r'<span[^>]*>([^<]+)</span>', ctx)
                if lm:
                    label = lm.group(1).strip()
            if not label:
                label = val
            radio_groups.setdefault(gid, []).append((val, label))

        # Évite les doublons radio (PrestaShop génère parfois une copie hidden)
        for gid, opts in radio_groups.items():
            seen = set()
            uniq = []
            for v, l in opts:
                if v not in seen:
                    seen.add(v)
                    uniq.append((v, l))
            if uniq:
                gname = group_id_to_name.get(gid, f"Group {gid}")
                # Évite de dupliquer si on a déjà ce groupe via select
                already_added = any(g[0] == gname for g in groups)
                if not already_added:
                    groups.append((gname, uniq))

        return groups

    @staticmethod
    def _extract_images(html_text: str, jsonld_images: list[dict]) -> list[dict]:
        """Récupère les URLs images. Combine plusieurs sources Presta 1.5/1.6/1.7+."""
        seen: list[str] = []
        # 1. JSON-LD si dispo
        for img in jsonld_images:
            src = img.get("src") if isinstance(img, dict) else None
            if src and src not in seen:
                seen.append(src)
        # 2. data-image-large-src (Presta 1.7+)
        for u in re.findall(r'data-image-large-src=["\']([^"\']+)["\']', html_text):
            if u not in seen:
                seen.append(u)
        # 3. .product-cover (Presta 1.7+)
        for u in re.findall(
            r'<img[^>]+class="[^"]*(?:product-cover|js-qv-product-cover)[^"]*"[^>]+src=["\']([^"\']+)["\']',
            html_text,
        ):
            if u not in seen:
                seen.append(u)
        # 4. itemprop="image" (Presta legacy + standard schema.org)
        for u in re.findall(
            r'<img[^>]+itemprop="image"[^>]+src=["\']([^"\']+)["\']',
            html_text,
        ):
            if u not in seen:
                seen.append(u)
        # 5. #bigpic (Presta 1.5/1.6 legacy)
        for u in re.findall(
            r'<img[^>]+id="bigpic"[^>]+src=["\']([^"\']+)["\']',
            html_text,
        ):
            if u not in seen:
                seen.append(u)
        # 6. .thickbox liens (galerie Presta legacy)
        for u in re.findall(
            r'<a[^>]+class="[^"]*thickbox[^"]*"[^>]+href=["\']([^"\']+)["\']',
            html_text,
        ):
            if u not in seen:
                seen.append(u)
        return [{"src": u} for u in seen]

    @staticmethod
    def _extract_description(html_text: str, dp_desc: str = "", jsonld_desc: str = "") -> str:
        """Récupère la description longue.

        Ordre de priorité :
          1. data-product .description (souvent complet, déjà en HTML)
          2. section.product-description / #description du HTML
          3. JSON-LD description (en fallback)
        """
        if dp_desc and len(dp_desc) > len(jsonld_desc):
            return dp_desc
        patterns = [
            # section.product-description (englobe toute la zone description)
            r'<section[^>]+class="[^"]*product-description[^"]*"[^>]*>(.*?)</section>',
            # #description (id) — utilisé par certains thèmes
            r'<div[^>]+id=["\']description["\'][^>]*>(.*?)</div>\s*(?:<div|<section|</section)',
            # itemprop=description
            r'<div[^>]+itemprop=["\']description["\'][^>]*>(.*?)</div>',
        ]
        for p in patterns:
            m = re.search(p, html_text, flags=re.DOTALL | re.IGNORECASE)
            if m:
                return m.group(1)
        # Fallback JSON-LD
        return jsonld_desc

    @staticmethod
    def _is_category_or_listing_page(html_text: str) -> bool:
        """Détecte si la page est une page catégorie/listing (et non une fiche produit).

        Cas typique : Promolinge redirige les URLs de produits supprimés vers leur
        page catégorie parente, en HTTP 200 (au lieu de 404). On veut skip ça
        silencieusement, sans essayer d'extraire un produit.
        """
        # Body class signaling category / listing pages
        m = re.search(r'<body[^>]+class="([^"]+)"', html_text[:5000])
        if m:
            body_classes = m.group(1).lower()
            for marker in ("page-category", "page-listing", "page-search",
                           "category-id-", "listing-page", "page-index"):
                if marker in body_classes:
                    return True
        # JSON-LD ItemList without Product = listing
        if "ItemList" in html_text and "data-product=" not in html_text:
            return True
        return False

    def _extract_product_from_html(
        self, url: str, html_text: str,
    ) -> WooProduct | None:
        """Construit un WooProduct depuis une page produit PrestaShop."""
        # FAST-PATH : page catégorie/listing (souvent une redirection silencieuse
        # depuis une URL produit supprimée → on skip sans bruit)
        if self._is_category_or_listing_page(html_text):
            self.logger.info(
                f"[Presta] {url} : page catégorie/listing détectée "
                f"(URL produit obsolète redirigée), skip"
            )
            return None

        dp = self._extract_data_product(html_text)
        ld = self._extract_jsonld_product(html_text)

        # Fallback : ancien PrestaShop (1.5/1.6) sans data-product
        if not dp:
            dp = self._extract_legacy_presta_product(url, html_text)
            if dp:
                self.logger.debug(f"[Presta] {url} : extraction via legacy fallback")

        if not dp and not ld:
            # Debug : dump le HTML reçu pour qu'on inspecte ce qui ne va pas
            try:
                import os
                debug_dir = Path("debug_presta_fails")
                debug_dir.mkdir(exist_ok=True)
                fn = re.sub(r"[^a-zA-Z0-9._-]", "_", url.split("//")[-1])[:80] + ".html"
                fp = debug_dir / fn
                with open(fp, "w", encoding="utf-8") as f:
                    f.write(html_text)
                # Indices de diagnostic
                has_dp_str = "data-product=" in html_text
                has_ldjson_str = "application/ld+json" in html_text
                has_presta_str = "prestashop" in html_text.lower()
                self.logger.warning(
                    f"[Presta] {url} : extraction échouée. "
                    f"len={len(html_text)} dp_text={has_dp_str} "
                    f"ldjson_text={has_ldjson_str} presta={has_presta_str}. "
                    f"HTML dumpé : {fp}"
                )
            except Exception:
                self.logger.warning(f"[Presta] {url} : ni data-product ni JSON-LD trouvés")
            return None

        # Récupère les champs avec priorité data-product > JSON-LD
        def _get(field: str, default=None):
            if dp and field in dp:
                return dp[field]
            if ld and field in ld:
                return ld[field]
            return default

        # ID Presta
        try:
            product_id = int((dp or {}).get("id_product") or 0)
        except (ValueError, TypeError):
            product_id = 0
        if not product_id:
            # Fallback : extrait depuis l'URL /{cat}/{id}-{slug}.html
            m_id = re.search(r"/(\d+)(?:-\d+)?-[^/]+\.html?$", url)
            if m_id:
                product_id = int(m_id.group(1))

        # Champs principaux
        name = (dp or {}).get("name") or (ld.get("name") if ld else "") or ""
        name = html.unescape(name).strip()
        slug = (dp or {}).get("link_rewrite") or ""
        sku = ((dp or {}).get("reference") or (ld.get("sku") if ld else "") or "").strip()

        # Description : on prend du data-product en priorité (souvent en HTML), sinon HTML page
        dp_description = (dp or {}).get("description") or ""
        dp_short = (dp or {}).get("description_short") or ""
        ld_description = (ld.get("description") if ld else "") or ""
        full_desc_html = self._extract_description(html_text, dp_description, ld_description)

        # Catégorie
        category_name = (dp or {}).get("category_name") or (ld.get("category") if ld else "") or ""
        categories = [{"name": category_name}] if category_name else []

        # Prix : en priorité data-product.price_amount (float), sinon JSON-LD offers.price
        price_amount = (dp or {}).get("price_amount")
        if price_amount is None and ld:
            offers = ld.get("offers")
            if isinstance(offers, dict):
                price_amount = offers.get("price") or offers.get("lowPrice")
        try:
            price_float = float(price_amount) if price_amount is not None else None
        except (ValueError, TypeError):
            price_float = None
        # Convertit en centimes pour cohérence avec WooScraper
        parent_price_cents = str(int(round(price_float * 100))) if price_float else None

        # Currency
        currency = "EUR"
        if ld and isinstance(ld.get("offers"), dict):
            currency = ld["offers"].get("priceCurrency", "EUR")

        # Stock parent
        stock_raw = (dp or {}).get("quantity")
        try:
            stock_qty = int(stock_raw) if stock_raw is not None else None
        except (ValueError, TypeError):
            stock_qty = None
        availability = (dp or {}).get("availability") or ""
        in_stock = availability == "available" or (stock_qty is not None and stock_qty > 0)

        # Images : combine JSON-LD + data-image-large-src
        ld_images = []
        if ld:
            ld_img = ld.get("image")
            if isinstance(ld_img, list):
                ld_images = [{"src": x if isinstance(x, str) else x.get("url")} for x in ld_img]
            elif isinstance(ld_img, str):
                ld_images = [{"src": ld_img}]
            elif isinstance(ld_img, dict):
                ld_images = [{"src": ld_img.get("url")}]
        images = self._extract_images(html_text, ld_images)
        # Nettoie les None
        images = [i for i in images if i.get("src")]

        # Variantes : extraire groupes d'attributs depuis le form
        groups = self._extract_attribute_groups(html_text, dp)

        variations_data: list[dict] = []
        variations_ids: list[int] = []

        if groups:
            # Produit cartésien des combinaisons
            from itertools import product as iproduct
            group_options = [g[1] for g in groups]
            group_names = [g[0] for g in groups]
            combos = list(iproduct(*group_options))
            self.logger.debug(
                f"[Presta] {url} : {len(groups)} groupes, "
                f"{len(combos)} combinaisons générées"
            )
            for idx, combo in enumerate(combos):
                # Construit un ID synthétique stable
                var_id = product_id * 10000 + idx
                # Attributs : [{"name": "Dimensions", "value": "190x90cm"}, ...]
                attrs = []
                sku_parts = []
                for gname, (vid, vlabel) in zip(group_names, combo):
                    attrs.append({"name": gname, "value": vlabel})
                    sku_parts.append(vid)
                # SKU : combine parent sku + IDs des combos
                v_sku = f"{sku or product_id}-{'-'.join(sku_parts)}" if sku_parts else ""
                variations_ids.append(var_id)
                variations_data.append({
                    "id": var_id,
                    "sku": v_sku,
                    "prices": {
                        "price": parent_price_cents,
                        "regular_price": parent_price_cents,
                        "currency_code": currency,
                        "currency_minor_unit": 2,
                    },
                    # On reporte le stock parent sur chaque combo (pas de moyen
                    # simple d'avoir le stock par combo sans appel AJAX)
                    "is_in_stock": in_stock,
                    "low_stock_remaining": None,
                    "images": [],  # variant images = pas dans le HTML initial
                    "_parent_attributes": attrs,
                })
            p_type = "variable"
        else:
            p_type = "simple"

        # Features : extraire et stocker tel quel (routing fait au moment du mapping)
        features_raw = (dp or {}).get("features") or []
        features = [
            {"name": str(f.get("name") or ""), "value": str(f.get("value") or "")}
            for f in features_raw
            if isinstance(f, dict) and f.get("name") and f.get("value")
        ]

        return WooProduct(
            id=product_id,
            name=name,
            slug=slug,
            type=p_type,
            sku=sku,
            short_description=dp_short or ld_description,
            description=full_desc_html,
            permalink=url,
            prices={
                "price": parent_price_cents,
                "currency_code": currency,
                "currency_minor_unit": 2,
            } if parent_price_cents else {},
            images=images,
            categories=categories,
            tags=[],
            attributes=[],
            variations_ids=variations_ids,
            variations_data=variations_data,
            is_in_stock=in_stock,
            low_stock_remaining=stock_qty,
            raw={"_source": "prestashop", "_url": url},
            features=features,
        )

    # -- Pipeline complet ------------------------------------------------

    def _fetch_and_extract_one(self, url: str) -> tuple[str, WooProduct | None, str]:
        """Worker function : fetch + extract pour une URL. Retourne (url, product, error_msg)."""
        try:
            status, hdrs, body = http_get_json(url, self.logger)
            if status != 200 or not isinstance(body, str):
                return url, None, f"status={status}"
            product = self._extract_product_from_html(url, body)
            return url, product, ""
        except Exception as e:
            return url, None, f"{type(e).__name__}: {e}"

    def build(self, max_products: int | None = None,
              concurrency: int = 2) -> list[WooProduct]:
        product_urls = self._get_product_urls()
        if not product_urls:
            raise HTTPError(f"PrestaShop : aucun sitemap exploitable sur {self.base}")

        self.logger.info(f"PrestaShop : {len(product_urls)} URLs produits à scraper")
        if max_products:
            product_urls = product_urls[:max_products]
            self.logger.info(f"  (limité à {len(product_urls)} pour ce run)")

        # Fetching concurrent. Default = 2 threads (compromis sécurité / vitesse).
        # Trop de threads écrase les petits serveurs (Promolinge → 100% timeouts à 4).
        # Use --concurrency 1 pour serveurs très lents, --concurrency 4-8 pour rapides.
        from concurrent.futures import ThreadPoolExecutor, as_completed
        self.logger.info(f"  concurrence: {concurrency} thread(s)")

        products: list[WooProduct] = []
        failures = 0
        timeouts_in_a_row = 0
        done = 0
        aborted = False

        # Si N timeouts consécutifs au DÉBUT sans aucun produit récupéré,
        # on abandonne : le serveur est probablement down. Évite de gaspiller
        # 5+ minutes sur un serveur injoignable.
        ABORT_AFTER_CONSECUTIVE_TIMEOUTS = 5

        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {executor.submit(self._fetch_and_extract_one, url): url
                       for url in product_urls}
            for fut in as_completed(futures):
                if aborted:
                    fut.cancel()
                    continue
                done += 1
                url, product, err = fut.result()
                if product:
                    products.append(product)
                    timeouts_in_a_row = 0
                else:
                    failures += 1
                    if err:
                        self.logger.warning(f"[Presta] {url} {err}")
                    if "timed out" in err.lower():
                        timeouts_in_a_row += 1
                    else:
                        timeouts_in_a_row = 0
                if done % 10 == 0 or done == len(product_urls):
                    self.logger.info(
                        f"  page {done}/{len(product_urls)} ({failures} échecs)"
                    )
                # Circuit-breaker : trop de timeouts consécutifs au début → abort
                if (timeouts_in_a_row >= ABORT_AFTER_CONSECUTIVE_TIMEOUTS
                        and len(products) == 0):
                    self.logger.error(
                        f"{timeouts_in_a_row} timeouts consécutifs sans aucun "
                        f"produit récupéré. Le serveur de cette marque est "
                        f"probablement down ou trop lent. Abandon du run — "
                        f"relance plus tard, ou avec --concurrency 1 si tu "
                        f"étais en concurrent."
                    )
                    aborted = True
                    executor.shutdown(wait=False, cancel_futures=True)
                    break

        # Si aborted, log clair sur ce qui s'est passé
        if aborted:
            raise HTTPError(
                f"PrestaShop {self.base} : abandon après "
                f"{timeouts_in_a_row} timeouts consécutifs (serveur down)"
            )
        if failures:
            self.logger.warning(f"PrestaShop terminé avec {failures}/{len(product_urls)} échecs")
        return products


# =============================================================================
# SECTION 6ter — Scraper Wix Stores (JSON-LD + Open Graph meta)
# =============================================================================

class WixScraper:
    """Scrape un catalogue Wix Stores via sitemap + HTML.

    Wix est JS-rendered : le contenu visible est chargé en AJAX. MAIS pour le
    SEO, Wix injecte côté serveur :
      - JSON-LD Schema.org Product (name, sku, brand, image, Offers, description)
      - Open Graph meta product:price:amount / product:price:currency
    Ces 2 sources suffisent à reconstruire un produit Ankorstore propre.

    Variantes : pas extractibles sans headless browser (chargées en AJAX). Pour
    la v1 on traite chaque page produit comme un produit simple.
    """

    def __init__(self, brand_url: str, logger: logging.Logger):
        self.brand_url = brand_url
        p = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
        self.base = f"{p.scheme}://{p.netloc}"
        self.domain = p.netloc.replace("www.", "")
        self.logger = logger

    # -- Récupération des URLs produits via sitemap index --------------

    def _get_product_urls(self) -> list[str]:
        """Wix utilise un sitemap.xml en index pointant vers store-products-sitemap.xml."""
        sm_index_url = f"{self.base}/sitemap.xml"
        try:
            status, hdrs, body = http_get_json(sm_index_url, self.logger)
        except HTTPError as e:
            self.logger.error(f"Sitemap index inaccessible : {e}")
            return []
        if status != 200 or not isinstance(body, str):
            self.logger.error(f"Sitemap index status={status}")
            return []

        # Le sitemap.xml de Wix est typiquement un index pointant vers
        # store-products-sitemap.xml, blog-posts-sitemap.xml, etc.
        sub_sitemaps = re.findall(r"<loc>([^<]+)</loc>", body)
        product_sitemap_url = None
        for sm in sub_sitemaps:
            if "store-products" in sm.lower() or "/products-sitemap" in sm.lower():
                product_sitemap_url = sm
                break
        if not product_sitemap_url:
            self.logger.warning(
                "Aucun sub-sitemap 'store-products' trouvé dans l'index. "
                "Cette boutique Wix n'a peut-être pas Wix Stores activé."
            )
            return []

        # Fetch le store-products-sitemap pour avoir les URLs produits
        try:
            status, hdrs, body = http_get_json(product_sitemap_url, self.logger)
        except HTTPError as e:
            self.logger.error(f"store-products-sitemap inaccessible : {e}")
            return []
        product_urls = re.findall(r"<loc>([^<]+)</loc>", body)
        # Dédoublonne en gardant l'ordre
        product_urls = list(dict.fromkeys(product_urls))
        self.logger.info(
            f"Wix : sitemap utilisé ({product_sitemap_url}), {len(product_urls)} URLs produits"
        )
        return product_urls

    # -- Extraction d'une page produit Wix ----------------------------

    @staticmethod
    def _extract_jsonld_product(html_text: str) -> dict | None:
        """Cherche un bloc JSON-LD de @type Product (cas typique sur Wix Stores)."""
        for block in re.findall(
            r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html_text, flags=re.DOTALL | re.IGNORECASE,
        ):
            try:
                data = json.loads(block.strip())
            except json.JSONDecodeError:
                continue
            candidates = data if isinstance(data, list) else (
                data.get("@graph", []) if isinstance(data, dict) and "@graph" in data else [data]
            )
            for c in candidates:
                if isinstance(c, dict):
                    t = c.get("@type")
                    if t == "Product" or (isinstance(t, list) and "Product" in t):
                        return c
        return None

    @staticmethod
    def _extract_og_meta(html_text: str) -> dict[str, str]:
        """Extrait les meta Open Graph (title/description/image/price/currency)."""
        out: dict[str, str] = {}
        for prop in ("og:title", "og:description", "og:image",
                     "product:price:amount", "product:price:currency",
                     "og:price:amount", "og:price:currency"):
            m = re.search(
                r'<meta[^>]+(?:property|name)=["\']'
                + re.escape(prop)
                + r'["\'][^>]+content=["\']([^"\']+)["\']',
                html_text, flags=re.IGNORECASE,
            )
            if m:
                out[prop] = html.unescape(m.group(1)).strip()
        return out

    @staticmethod
    def _extract_balanced_json_array(text: str, key: str) -> list | None:
        """Trouve la valeur d'un champ JSON `"key":[...]` et la parse.

        Utilise un matcher de crochets équilibrés (gère strings, escape, nested).
        Wix injecte un gros JSON dans le HTML qui contient productItems et options ;
        on ne peut pas utiliser regex naïve (nested objects). Cette fonction
        extrait le bloc [...] complet et json.loads.
        """
        pattern = re.escape(f'"{key}":') + r'\s*\['
        m = re.search(pattern, text)
        if not m:
            return None
        start = m.end() - 1  # position de [
        depth = 0
        in_string = False
        escape_next = False
        end = -1
        for i in range(start, len(text)):
            c = text[i]
            if escape_next:
                escape_next = False
                continue
            if c == '\\':
                escape_next = True
                continue
            if c == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break
        if end < 0:
            return None
        try:
            return json.loads(text[start:end])
        except json.JSONDecodeError:
            return None

    @staticmethod
    def _extract_wix_variants(html_text: str, currency: str = "EUR") -> list[dict]:
        """Reconstruit les variantes Wix depuis productItems + options.

        Retourne une liste de dicts compatibles avec le format `variations_data`
        utilisé par le mapper Ankorstore.
        """
        product_items = WixScraper._extract_balanced_json_array(
            html_text, "productItems",
        )
        options = WixScraper._extract_balanced_json_array(html_text, "options")
        if not product_items or not options:
            return []
        # Si un seul productItem, c'est un produit simple (pas de variantes)
        if len(product_items) <= 1:
            return []

        # Construit le mapping selection_id -> {"name": option_title, "value": ...}
        selection_to_attr: dict[int, dict] = {}
        for opt in options:
            if not isinstance(opt, dict):
                continue
            opt_title = opt.get("title") or opt.get("key") or "Option"
            for sel in opt.get("selections", []):
                if isinstance(sel, dict) and "id" in sel and "value" in sel:
                    selection_to_attr[sel["id"]] = {
                        "name": opt_title,
                        "value": sel["value"],
                    }

        variations: list[dict] = []
        for idx, item in enumerate(product_items):
            if not isinstance(item, dict):
                continue
            item_id = str(item.get("id", "") or "").strip()
            price = item.get("price")
            try:
                price_cents = str(int(round(float(price) * 100))) if price else None
            except (ValueError, TypeError):
                price_cents = None
            sels = item.get("optionsSelections", []) or []
            attrs = []
            for sel_id in sels:
                a = selection_to_attr.get(sel_id)
                if a:
                    attrs.append(a)
            # is_in_stock : Wix expose parfois inventory dans productItems
            inv = item.get("inventory") or {}
            in_stock = True
            if isinstance(inv, dict) and "status" in inv:
                in_stock = inv["status"].lower() == "in_stock"
            variations.append({
                "id": idx,
                "sku": item_id,
                "prices": {
                    "price": price_cents,
                    "regular_price": price_cents,
                    "currency_code": currency,
                    "currency_minor_unit": 2,
                },
                "is_in_stock": in_stock,
                "low_stock_remaining": None,
                "images": [],
                "_parent_attributes": attrs,
            })
        return variations

    @staticmethod
    def _jsonld_images_to_urls(image_field: Any) -> list[str]:
        """Normalise le champ image du JSON-LD vers une liste d'URLs."""
        if not image_field:
            return []
        items = image_field if isinstance(image_field, list) else [image_field]
        out = []
        for it in items:
            if isinstance(it, str):
                out.append(it)
            elif isinstance(it, dict):
                url = (it.get("contentUrl") or it.get("url")
                       or it.get("src") or "")
                if url:
                    out.append(url)
        return out

    def _extract_product_from_html(
        self, url: str, html_text: str,
    ) -> WooProduct | None:
        """Construit un WooProduct depuis une page produit Wix."""
        ld = self._extract_jsonld_product(html_text)
        og = self._extract_og_meta(html_text)

        if not ld and not og:
            self.logger.warning(
                f"[Wix] {url} : ni JSON-LD Product ni Open Graph meta trouvés"
            )
            return None

        # Nom : JSON-LD prioritaire, sinon og:title
        name = ""
        if ld:
            name = html.unescape(ld.get("name") or "").strip()
        if not name and og.get("og:title"):
            # og:title contient souvent " | Brand" — on strippera après
            name = og["og:title"]

        # Description : prend la plus longue entre JSON-LD et og:description
        ld_desc = html.unescape(ld.get("description") if ld else "").strip() if ld else ""
        og_desc = og.get("og:description", "")
        # JSON-LD description peut être "SKU: XXX" (rien d'utile) → préférer og dans ce cas
        if ld_desc.lower().startswith("sku:") or len(ld_desc) < 30:
            description = og_desc or ld_desc
        else:
            description = ld_desc if len(ld_desc) >= len(og_desc) else og_desc

        # SKU
        sku = (ld.get("sku") if ld else "") or ""
        sku = str(sku).strip()

        # Brand (peut être string ou {"@type":"Brand","name":"..."})
        brand_name = ""
        brand_field = (ld.get("brand") if ld else None)
        if isinstance(brand_field, dict):
            brand_name = brand_field.get("name") or ""
        elif isinstance(brand_field, str):
            brand_name = brand_field

        # Images
        images_urls = []
        if ld:
            images_urls = self._jsonld_images_to_urls(ld.get("image"))
        if not images_urls and og.get("og:image"):
            images_urls = [og["og:image"]]
        # Upscale Wix images (w_500 → w_2000)
        images_urls = [upscale_image_url(u) for u in images_urls]
        images = [{"src": u} for u in images_urls if u]

        # Prix : OG meta en priorité (très fiable sur Wix), sinon JSON-LD Offers
        price_amount = None
        currency = "EUR"
        og_price = og.get("product:price:amount") or og.get("og:price:amount")
        og_currency = og.get("product:price:currency") or og.get("og:price:currency")
        if og_price:
            try:
                price_amount = float(og_price)
            except ValueError:
                pass
        if og_currency:
            currency = og_currency
        # Fallback JSON-LD Offers (clé "Offers" avec O majuscule sur Wix !)
        if price_amount is None and ld:
            offers = ld.get("Offers") or ld.get("offers")
            if isinstance(offers, dict):
                try:
                    price_amount = float(offers.get("price") or offers.get("lowPrice") or 0) or None
                except (ValueError, TypeError):
                    pass
                if offers.get("priceCurrency"):
                    currency = offers["priceCurrency"]

        parent_price_cents = (
            str(int(round(price_amount * 100))) if price_amount else None
        )

        # ID synthétique stable depuis le slug
        slug_match = re.search(r"/(?:product-page|products)/([^/?#]+)", url)
        slug = slug_match.group(1) if slug_match else url.rstrip("/").rsplit("/", 1)[-1]
        product_id = abs(hash(slug)) % (10 ** 9)

        # Catégorie : pas exposée fiablement par Wix dans JSON-LD ; brand→category
        categories = []
        if brand_name:
            categories = [{"name": brand_name}]

        # Extraction des variantes depuis productItems + options JSON inline.
        # Si plusieurs productItems → produit variable, on génère 1 row par variante.
        variations_data = self._extract_wix_variants(html_text, currency)
        variations_ids = [v["id"] for v in variations_data]
        p_type = "variable" if len(variations_data) > 1 else "simple"
        if variations_data:
            self.logger.debug(
                f"[Wix] {url} : {len(variations_data)} variantes extraites"
            )

        return WooProduct(
            id=product_id,
            name=name,
            slug=slug,
            type=p_type,
            sku=sku,
            short_description=description,
            description="",
            permalink=url,
            prices={
                "price": parent_price_cents,
                "currency_code": currency,
                "currency_minor_unit": 2,
            } if parent_price_cents else {},
            images=images,
            categories=categories,
            tags=[],
            attributes=[],
            variations_ids=variations_ids,
            variations_data=variations_data,
            is_in_stock=True,  # Wix ne nous expose pas le stock global fiablement
            low_stock_remaining=None,
            raw={"_source": "wix", "_url": url},
            features=[],
        )

    # -- Worker thread ------------------------------------------------

    def _fetch_and_extract_one(self, url: str) -> tuple[str, WooProduct | None, str]:
        try:
            # Encode l'URL pour gérer les accents (ex: "pâté" devient "p%C3%A2t%C3%A9")
            safe_url = self._url_encode(url)
            status, hdrs, body = http_get_json(safe_url, self.logger)
            if status != 200 or not isinstance(body, str):
                return url, None, f"status={status}"
            product = self._extract_product_from_html(url, body)
            return url, product, ""
        except Exception as e:
            return url, None, f"{type(e).__name__}: {e}"

    @staticmethod
    def _url_encode(url: str) -> str:
        """Encode les caractères non-ASCII dans le path pour éviter les
        UnicodeEncodeError dans urllib."""
        from urllib.parse import quote, urlsplit, urlunsplit
        parts = urlsplit(url)
        new_path = quote(parts.path, safe="/-._~%")
        return urlunsplit((parts.scheme, parts.netloc, new_path, parts.query, parts.fragment))

    # -- Build pipeline -----------------------------------------------

    def build(self, max_products: int | None = None,
              concurrency: int = 2) -> list[WooProduct]:
        product_urls = self._get_product_urls()
        if not product_urls:
            raise HTTPError(f"Wix : aucune URL produit trouvée sur {self.base}")

        if max_products:
            product_urls = product_urls[:max_products]
            self.logger.info(f"  (limité à {len(product_urls)} pour ce run)")

        self.logger.info(f"  concurrence: {concurrency} thread(s)")

        from concurrent.futures import ThreadPoolExecutor, as_completed
        products: list[WooProduct] = []
        failures = 0
        done = 0

        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {executor.submit(self._fetch_and_extract_one, url): url
                       for url in product_urls}
            for fut in as_completed(futures):
                done += 1
                url, product, err = fut.result()
                if product:
                    products.append(product)
                else:
                    failures += 1
                    if err:
                        self.logger.warning(f"[Wix] {url} {err}")
                if done % 10 == 0 or done == len(product_urls):
                    self.logger.info(
                        f"  page {done}/{len(product_urls)} ({failures} échecs)"
                    )

        if failures:
            self.logger.warning(f"Wix terminé avec {failures}/{len(product_urls)} échecs")
        return products


# =============================================================================
# SECTION 6quater — Scraper Squarespace (?format=json natif)
# =============================================================================

class SquarespaceScraper:
    """Scrape un catalogue Squarespace via le trick `?format=json` natif.

    Squarespace renvoie le JSON brut d'une page si on ajoute `?format=json`.
    Pour une URL produit (/shop/p/{slug}), on récupère un objet `item` complet
    avec ses variants, prix, images, description HTML, tags, catégories.

    Stratégie :
      1. Sitemap → URLs produits (pattern /shop/p/{slug}, /products/{slug}, etc.)
      2. Pour chaque URL : GET ?format=json → parse → extract item
      3. Pour les variants : item.structuredContent.variants[] (price en centimes,
         attributes en dict, sku, image, stock)
    """

    def __init__(self, brand_url: str, logger: logging.Logger):
        self.brand_url = brand_url
        p = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
        self.base = f"{p.scheme}://{p.netloc}"
        self.domain = p.netloc.replace("www.", "")
        self.logger = logger

    # -- Découverte des URLs produits ----------------------------------

    @staticmethod
    def _looks_like_product_url(url: str) -> bool:
        """URL produit individuelle Squarespace : doit contenir `/p/{slug}`.

        Pattern Squarespace : `/shop/{collection?}/p/{product-slug}`. La présence
        du segment `/p/` est le marqueur fiable pour distinguer un produit d'une
        page collection (genre `/shop/bouquet` qui liste les produits).
        """
        return bool(re.search(
            r"/(?:shop|store|boutique|produits|products)(?:/[^/]+)*/p/[a-zA-Z0-9][a-zA-Z0-9\-_]+/?$",
            url,
        ))

    @staticmethod
    def _looks_like_collection_url(url: str) -> bool:
        """URL catégorie/collection Squarespace : /shop/{category} (sans /p/)."""
        # /shop/{x} où x ne contient pas /p/ avant
        if "/shop/p/" in url or "/products/p/" in url:
            return False
        return bool(re.search(r"/(?:shop|store|boutique|produits|products)/[^/]+/?$", url))

    def _get_product_urls(self) -> list[str]:
        """Découvre les URLs produits selon le contexte :

        MODE 1 (URL non-racine) : la brand_url EST une page collection
          ex: https://www.maisonfelicia.com/concept-store/selection
          → on fetch ?format=json directement sur cette URL pour extraire ses items

        MODE 2 (URL home) : on utilise le sitemap.xml
          → on filtre les /shop/p/{slug} (fiches produits individuelles)
          → fallback : on crawl les pages collections trouvées dans le sitemap
        """
        parsed = urlparse(self.brand_url)
        path = parsed.path.rstrip("/")
        is_root = (not path or path == "")

        # MODE 1 : URL collection scopée
        if not is_root and not self._looks_like_product_url(self.brand_url):
            self.logger.info(
                f"Squarespace : URL non-racine détectée → mode collection scoped "
                f"({self.brand_url})"
            )
            urls = self._crawl_one_collection(self.brand_url)
            if urls:
                self.logger.info(
                    f"  → {len(urls)} produits trouvés dans cette collection"
                )
                return urls
            self.logger.warning(
                "  → 0 produit trouvé dans cette collection, fallback sur sitemap"
            )

        # MODE 2 : sitemap
        try:
            s, h, body = http_get_json(f"{self.base}/sitemap.xml", self.logger)
        except HTTPError as e:
            raise HTTPError(f"sitemap.xml inaccessible : {e}")
        if s != 200 or not isinstance(body, str):
            raise HTTPError(f"sitemap.xml inaccessible (status={s})")
        all_locs = re.findall(r"<loc>([^<]+)</loc>", body)

        # 2a. URLs /shop/p/{slug} = fiches produits directes
        product_urls = [u for u in all_locs if self._looks_like_product_url(u)]
        if product_urls:
            self.logger.info(
                f"Squarespace : sitemap utilisé, {len(product_urls)} URLs produits "
                f"directes (format /shop/p/{{slug}})"
            )
            return list(dict.fromkeys(product_urls))

        # 2b. Fallback : crawl des pages collections du sitemap
        collection_urls = [u for u in all_locs if self._looks_like_collection_url(u)]
        if collection_urls:
            self.logger.info(
                f"Squarespace : pas de /shop/p/ dans sitemap, on crawl "
                f"{len(collection_urls)} pages collections du sitemap"
            )
            products: list[str] = []
            seen: set[str] = set()
            for col_url in collection_urls:
                col_products = self._crawl_one_collection(col_url)
                for p_url in col_products:
                    if p_url not in seen:
                        seen.add(p_url)
                        products.append(p_url)
                time.sleep(0.3)
            self.logger.info(f"  → {len(products)} produits récupérés via collections")
            return products

        return []

    def _crawl_one_collection(self, col_url: str) -> list[str]:
        """Pour une URL collection donnée, fetch ?format=json et extrait les
        URLs des produits trouvés dans items[]."""
        col_json = self._fetch_format_json(col_url)
        if not col_json:
            return []
        products: list[str] = []
        for it in (col_json.get("items") or []):
            if not isinstance(it, dict):
                continue
            full_url = it.get("fullUrl") or ""
            if full_url:
                p_url = self.base + full_url if full_url.startswith("/") else full_url
                products.append(p_url)
        # Dédoublonne en gardant l'ordre
        return list(dict.fromkeys(products))

    # -- Fetch JSON natif Squarespace ---------------------------------

    def _fetch_format_json(self, url: str) -> dict | None:
        json_url = url + ("&" if "?" in url else "?") + "format=json"
        try:
            s, h, body = http_get_json(json_url, self.logger)
        except HTTPError:
            return None
        if s != 200:
            return None
        # Squarespace renvoie JSON ; si on a déjà un dict, c'est parse fait
        if isinstance(body, dict):
            return body
        if isinstance(body, str):
            try:
                return json.loads(body)
            except json.JSONDecodeError:
                return None
        return None

    # -- Construction du WooProduct depuis le JSON Squarespace --------

    def _extract_product(self, url: str, data: dict) -> WooProduct | None:
        """Construit un WooProduct depuis le JSON ?format=json d'une page produit."""
        item = data.get("item")
        if not isinstance(item, dict):
            self.logger.warning(f"[Squarespace] {url} : pas d'objet 'item' dans le JSON")
            return None

        # ID synthétique stable depuis le slug
        slug = (item.get("urlId") or item.get("filename") or
                url.rstrip("/").rsplit("/", 1)[-1])
        try:
            product_id = int(item.get("id") or 0)
        except (ValueError, TypeError):
            product_id = abs(hash(slug)) % (10 ** 9)

        # Nom & description
        name = html.unescape(item.get("title") or "").strip()
        # Squarespace stocke la description soit dans `body` (long, rich text),
        # soit dans `excerpt` (short description). Beaucoup de marques (cas
        # Maison Felicia) utilisent uniquement excerpt → on fallback.
        body_text = (item.get("body") or "").strip()
        excerpt = (item.get("excerpt") or "").strip()
        if body_text and excerpt and excerpt not in body_text:
            # Les deux remplis ET différents → on concatène
            body_html = body_text + "\n\n" + excerpt
        else:
            body_html = body_text or excerpt

        # Tags & catégories
        tags = [{"name": t} for t in (item.get("tags") or [])]
        categories = [{"name": c} for c in (item.get("categories") or [])]

        # Variants Squarespace : structuredContent.variants ou items
        sc = item.get("structuredContent") or {}
        variants_raw = sc.get("variants") or item.get("items") or []
        if not isinstance(variants_raw, list):
            variants_raw = []

        # Currency (souvent dans websiteSettings)
        currency = "EUR"
        ws_settings = data.get("websiteSettings") or {}
        store_settings = ws_settings.get("storeSettings") or {}
        if store_settings.get("currency"):
            currency = store_settings["currency"]

        # Construit la liste des variantes au format compatible mapper
        variations_data = []
        for idx, v in enumerate(variants_raw):
            if not isinstance(v, dict):
                continue
            v_sku = (v.get("sku") or "").strip()
            # Squarespace prix en centimes (price=5000 → 50.00€) — déjà ce qu'on veut
            v_price = v.get("price")
            v_sale = v.get("salePrice")
            # On préfère le sale price s'il est défini et > 0
            effective_price = v_sale if (v_sale and v_sale > 0) else v_price
            v_attrs = []
            attrs_obj = v.get("attributes") or {}
            if isinstance(attrs_obj, dict):
                for k, val in attrs_obj.items():
                    if val:
                        v_attrs.append({"name": k, "value": str(val)})
            # Stock : -1 = unlimited dans Squarespace
            stock_raw = v.get("stock")
            v_in_stock = True
            v_stock_qty: int | None = None
            if isinstance(stock_raw, int):
                if stock_raw == -1:
                    v_in_stock = True
                    v_stock_qty = None
                elif stock_raw == 0:
                    v_in_stock = False
                    v_stock_qty = 0
                else:
                    v_in_stock = True
                    v_stock_qty = stock_raw
            # Image variant
            v_imgs = []
            img_field = v.get("image")
            if isinstance(img_field, dict):
                src = (img_field.get("assetUrl") or img_field.get("url") or "")
                if src:
                    v_imgs.append({"src": src})
            variations_data.append({
                "id": idx,
                "sku": v_sku,
                "prices": {
                    "price": str(effective_price) if effective_price else None,
                    "regular_price": str(v_price) if v_price else None,
                    "currency_code": currency,
                    "currency_minor_unit": 2,
                },
                "is_in_stock": v_in_stock,
                "low_stock_remaining": v_stock_qty,
                "images": v_imgs,
                "_parent_attributes": v_attrs,
            })

        # Type & prix parent
        p_type = "variable" if len(variations_data) > 1 else "simple"
        parent_price_cents = None
        if variations_data:
            # Prix parent = prix de la 1ère variante
            first_price = variations_data[0]["prices"]["price"]
            parent_price_cents = first_price

        # Images parent : on combine main image + tous les items.image
        images_urls: list[str] = []
        asset = item.get("assetUrl")
        if asset:
            images_urls.append(asset)
        # additionalImages dans structuredContent
        for sc_img in (sc.get("images") or item.get("images") or []):
            if isinstance(sc_img, dict):
                src = sc_img.get("assetUrl") or sc_img.get("url") or ""
                if src and src not in images_urls:
                    images_urls.append(src)
        # Images des variantes (uniques)
        for v in variations_data:
            for img in v.get("images") or []:
                src = img.get("src", "")
                if src and src not in images_urls:
                    images_urls.append(src)
        images = [{"src": u} for u in images_urls]

        # Stock parent
        in_stock = any(v["is_in_stock"] for v in variations_data) if variations_data else True

        return WooProduct(
            id=product_id,
            name=name,
            slug=slug,
            type=p_type,
            sku=(item.get("sku") or "").strip(),
            short_description=body_html,
            description="",
            permalink=url,
            prices={
                "price": parent_price_cents,
                "currency_code": currency,
                "currency_minor_unit": 2,
            } if parent_price_cents else {},
            images=images,
            categories=categories,
            tags=tags,
            attributes=[],
            variations_ids=[v["id"] for v in variations_data],
            variations_data=variations_data,
            is_in_stock=in_stock,
            low_stock_remaining=None,
            raw={"_source": "squarespace", "_url": url},
            features=[],
        )

    # -- Worker ---------------------------------------------------------

    def _fetch_and_extract_one(self, url: str) -> tuple[str, WooProduct | None, str]:
        try:
            data = self._fetch_format_json(url)
            if not data:
                return url, None, "fetch_format_json a renvoyé vide"
            product = self._extract_product(url, data)
            return url, product, ""
        except Exception as e:
            return url, None, f"{type(e).__name__}: {e}"

    # -- Pipeline -------------------------------------------------------

    def build(self, max_products: int | None = None,
              concurrency: int = 2) -> list[WooProduct]:
        product_urls = self._get_product_urls()
        if not product_urls:
            raise HTTPError(
                f"Squarespace : aucune URL produit trouvée sur {self.base}. "
                f"Cette marque n'a peut-être pas Squarespace Stores activé."
            )

        if max_products:
            product_urls = product_urls[:max_products]
            self.logger.info(f"  (limité à {len(product_urls)} pour ce run)")
        self.logger.info(f"  concurrence: {concurrency} thread(s)")

        from concurrent.futures import ThreadPoolExecutor, as_completed
        products: list[WooProduct] = []
        failures = 0
        done = 0
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {executor.submit(self._fetch_and_extract_one, url): url
                       for url in product_urls}
            for fut in as_completed(futures):
                done += 1
                url, product, err = fut.result()
                if product:
                    products.append(product)
                else:
                    failures += 1
                    if err:
                        self.logger.warning(f"[Squarespace] {url} {err}")
                if done % 10 == 0 or done == len(product_urls):
                    self.logger.info(f"  page {done}/{len(product_urls)} ({failures} échecs)")
        if failures:
            self.logger.warning(f"Squarespace terminé avec {failures}/{len(product_urls)} échecs")
        return products


# =============================================================================
# SECTION 6quinquies — Scraper Custom (sites sans CMS connu)
# =============================================================================

class CustomScraper:
    """Scrape un site e-commerce sans CMS connu, en exploitant les signaux SEO
    universels présents sur la plupart des sites :
      1. JSON-LD Schema.org Product (idéal — quand présent)
      2. Open Graph meta product:price:amount/currency, og:title/description/image
      3. Microdata itemprop="price"/"name"/"description"/"sku" (souvent dans
         l'attribut `content="..."` plutôt que dans le texte)

    Stratégie URLs produits :
      - Sitemap si dispo, filtre URLs avec pattern produit-like
      - Sinon crawl basique de la home

    Reste un best-effort : selon les marques, qualité variable. Ne supporte pas
    les variantes (chaque page = 1 produit simple).
    """

    def __init__(self, brand_url: str, logger: logging.Logger):
        self.brand_url = brand_url
        p = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
        self.base = f"{p.scheme}://{p.netloc}"
        self.domain = p.netloc.replace("www.", "")
        self.logger = logger

    # -- Découverte URLs --------------------------------------------------

    @staticmethod
    def _looks_like_product_url(url: str) -> bool:
        """Heuristiques pour détecter une URL produit."""
        path = urlparse(url).path
        # /produit/X, /product/X, /shop/X, /boutique/X
        if re.search(r"/(?:product|produit|shop|store|boutique|p)/[a-z0-9]", path, re.I):
            return True
        # /{N}-{slug} (cas lessenteursdelanature)
        if re.search(r"/\d{4,}-[a-zA-Z0-9\-_]+/?$", path):
            return True
        # /{slug}.html (cas Presta-like sans CMS détecté)
        if re.search(r"/[a-z0-9][a-z0-9\-_]{4,}\.html?$", path, re.I):
            return True
        return False

    def _get_product_urls(self) -> list[str]:
        parsed = urlparse(self.brand_url)
        path = parsed.path.rstrip("/")
        is_root = (not path or path == "")

        # MODE DIRECT : URL non-racine = collection ou produit individuel scopé
        if not is_root:
            if self._looks_like_product_url(self.brand_url):
                self.logger.info(f"Mode produit unique : {self.brand_url}")
                return [self.brand_url]
            self.logger.info(
                f"Mode collection/page scopée : {self.brand_url} — on crawl pour "
                f"trouver les produits"
            )
            return self._crawl_for_product_urls_from(self.brand_url)

        # MODE SITEMAP
        try:
            s, hdrs, body = http_get_json(f"{self.base}/sitemap.xml", self.logger)
        except HTTPError:
            s, body = 0, ""
        if s == 200 and isinstance(body, str) and "<loc>" in body:
            locs = re.findall(r"<loc>([^<]+)</loc>", body)
            # Sitemap index ? suivre les sub-sitemaps
            if "<sitemapindex" in body[:300]:
                all_locs: list[str] = []
                for sub in locs:
                    try:
                        s2, _, b2 = http_get_json(sub, self.logger)
                        if s2 == 200 and isinstance(b2, str):
                            all_locs.extend(re.findall(r"<loc>([^<]+)</loc>", b2))
                    except HTTPError:
                        continue
                    time.sleep(0.3)
                locs = all_locs
            products = [u for u in locs
                        if self.domain in u and self._looks_like_product_url(u)]
            if products:
                self.logger.info(f"Custom : sitemap utilisé, {len(products)} URLs produits")
                return list(dict.fromkeys(products))

        # FALLBACK : crawl home pour trouver des liens produits
        self.logger.warning("Sitemap inexploitable, crawl de la home en fallback")
        return self._crawl_for_product_urls_from(self.brand_url)

    def _crawl_for_product_urls_from(self, page_url: str) -> list[str]:
        """Crawl une page et extrait toutes les URLs internes qui ressemblent à
        des produits."""
        try:
            s, h, body = http_get_json(page_url, self.logger)
        except HTTPError:
            return []
        if s != 200 or not isinstance(body, str):
            return []
        product_urls: list[str] = []
        seen: set[str] = set()
        for href in re.findall(r'href=["\']([^"\']+)["\']', body):
            if href.startswith("//"):
                href = "https:" + href
            elif href.startswith("/"):
                href = self.base + href
            elif not href.startswith("http"):
                continue
            if self.domain not in href:
                continue
            href_clean = href.split("#")[0].split("?")[0]
            if href_clean in seen:
                continue
            if self._looks_like_product_url(href_clean):
                seen.add(href_clean)
                product_urls.append(href_clean)
        return product_urls

    # -- Extraction d'une page produit -----------------------------------

    @staticmethod
    def _extract_jsonld_product(html_text: str) -> dict | None:
        for block in re.findall(
            r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html_text, flags=re.DOTALL | re.IGNORECASE,
        ):
            try:
                data = json.loads(block.strip())
            except json.JSONDecodeError:
                continue
            candidates = data if isinstance(data, list) else (
                data.get("@graph", []) if isinstance(data, dict) and "@graph" in data else [data]
            )
            for c in candidates:
                if isinstance(c, dict):
                    t = c.get("@type")
                    if t == "Product" or (isinstance(t, list) and "Product" in t):
                        return c
        return None

    @staticmethod
    def _extract_og_meta(html_text: str) -> dict[str, str]:
        out: dict[str, str] = {}
        for prop in ("og:title", "og:description", "og:image",
                     "product:price:amount", "product:price:currency",
                     "og:price:amount", "og:price:currency",
                     "twitter:title", "twitter:description"):
            m = re.search(
                r'<meta[^>]+(?:property|name)=["\']' + re.escape(prop) + r'["\'][^>]+content=["\']([^"\']+)["\']',
                html_text, flags=re.IGNORECASE,
            )
            if m:
                out[prop] = html.unescape(m.group(1)).strip()
        return out

    @staticmethod
    def _extract_microdata(html_text: str) -> dict[str, str]:
        """Extrait microdata itemprop. Cherche dans l'attribut content="..." puis
        dans le texte de l'élément, prend ce qui est non-vide."""
        out: dict[str, str] = {}
        for prop in ("price", "priceCurrency", "name", "description", "sku",
                     "availability", "image", "brand"):
            # Pattern 1 : content="..." dans n'importe quel ordre d'attribut
            value = None
            for pat in (
                rf'<[^>]*itemprop=["\']{prop}["\'][^>]*content=["\']([^"\']+)["\']',
                rf'<[^>]*content=["\']([^"\']+)["\'][^>]*itemprop=["\']{prop}["\']',
            ):
                m = re.search(pat, html_text, flags=re.IGNORECASE | re.DOTALL)
                if m:
                    value = m.group(1).strip()
                    if value:
                        break
            if not value:
                # Pattern 2 : contenu texte entre <...itemprop="X">TEXTE</...>
                m = re.search(
                    rf'<[^>]+itemprop=["\']{prop}["\'][^>]*>([^<]+)<',
                    html_text, flags=re.IGNORECASE,
                )
                if m:
                    candidate = m.group(1).strip()
                    if candidate and len(candidate) > 0:
                        value = candidate
            if value:
                out[prop] = html.unescape(value)
        return out

    def _extract_product_from_html(self, url: str, body: str) -> WooProduct | None:
        ld = self._extract_jsonld_product(body)
        og = self._extract_og_meta(body)
        md = self._extract_microdata(body)

        if not ld and not og and not md:
            self.logger.warning(f"[Custom] {url} : aucune donnée structurée trouvée")
            return None

        # Nom
        name = ""
        if ld:
            name = html.unescape(ld.get("name") or "").strip()
        if not name and og.get("og:title"):
            name = og["og:title"]
        if not name and md.get("name"):
            name = md["name"]

        # Description : cherche la plus longue parmi les sources
        desc_candidates = []
        if ld and ld.get("description"):
            desc_candidates.append(html.unescape(ld["description"]).strip())
        if og.get("og:description"):
            desc_candidates.append(og["og:description"])
        if md.get("description"):
            desc_candidates.append(md["description"])
        description = max(desc_candidates, key=len) if desc_candidates else ""

        # SKU : depuis JSON-LD, microdata, sinon fallback sur URL/ID
        sku = ""
        if ld and ld.get("sku"):
            sku = str(ld["sku"]).strip()
        if not sku and md.get("sku"):
            sku = md["sku"]

        # Brand
        brand_name = ""
        if ld:
            b = ld.get("brand")
            if isinstance(b, dict):
                brand_name = b.get("name") or ""
            elif isinstance(b, str):
                brand_name = b
        if not brand_name and md.get("brand"):
            brand_name = md["brand"]

        # Prix
        price_amount = None
        currency = "EUR"
        # 1. OG meta (souvent fiable)
        og_price = og.get("product:price:amount") or og.get("og:price:amount")
        if og_price:
            try:
                price_amount = float(og_price)
            except ValueError:
                pass
            currency = og.get("product:price:currency") or og.get("og:price:currency") or currency
        # 2. Microdata
        if price_amount is None and md.get("price"):
            try:
                # microdata price can be "17" or "17,00" or "17.00"
                price_str = md["price"].replace(",", ".")
                price_amount = float(price_str)
            except (ValueError, TypeError):
                pass
            md_currency = md.get("priceCurrency")
            if md_currency:
                # priceCurrency peut être "EUR" ou "€" → normalise
                if md_currency.upper() == "EUR" or md_currency == "€":
                    currency = "EUR"
                elif md_currency.upper() in ("USD", "GBP", "CHF"):
                    currency = md_currency.upper()
        # 3. JSON-LD Offers
        if price_amount is None and ld:
            offers = ld.get("offers")
            if isinstance(offers, dict):
                try:
                    price_amount = float(offers.get("price") or offers.get("lowPrice") or 0) or None
                except (ValueError, TypeError):
                    pass
                if offers.get("priceCurrency"):
                    currency = offers["priceCurrency"]

        parent_price_cents = (
            str(int(round(price_amount * 100))) if price_amount else None
        )

        # Images
        images_urls: list[str] = []
        if ld:
            ld_img = ld.get("image")
            if isinstance(ld_img, list):
                for x in ld_img:
                    if isinstance(x, str):
                        images_urls.append(x)
                    elif isinstance(x, dict):
                        url_str = x.get("url") or x.get("contentUrl") or x.get("src")
                        if url_str:
                            images_urls.append(url_str)
            elif isinstance(ld_img, str):
                images_urls.append(ld_img)
            elif isinstance(ld_img, dict):
                url_str = ld_img.get("url") or ld_img.get("contentUrl") or ld_img.get("src")
                if url_str:
                    images_urls.append(url_str)
        if og.get("og:image") and og["og:image"] not in images_urls:
            images_urls.append(og["og:image"])
        if md.get("image") and md["image"] not in images_urls:
            images_urls.append(md["image"])
        images = [{"src": upscale_image_url(u)} for u in images_urls if u]

        # Categorie via brand
        categories = [{"name": brand_name}] if brand_name else []

        # ID synthétique stable
        path = urlparse(url).path
        m_id = re.search(r"/(\d+)-", path)
        if m_id:
            product_id = int(m_id.group(1))
        else:
            product_id = abs(hash(path)) % (10 ** 9)

        # SKU fallback : ID dans l'URL
        if not sku and m_id:
            sku = m_id.group(1)

        # Stock : on n'a pas l'info → assume in_stock
        in_stock = True
        if md.get("availability"):
            avail = md["availability"].lower()
            if "outofstock" in avail or "out_of_stock" in avail:
                in_stock = False

        return WooProduct(
            id=product_id,
            name=name,
            slug=path.strip("/").replace("/", "-"),
            type="simple",
            sku=sku,
            short_description=description,
            description="",
            permalink=url,
            prices={
                "price": parent_price_cents,
                "currency_code": currency,
                "currency_minor_unit": 2,
            } if parent_price_cents else {},
            images=images,
            categories=categories,
            tags=[],
            attributes=[],
            variations_ids=[],
            variations_data=[],
            is_in_stock=in_stock,
            low_stock_remaining=None,
            raw={"_source": "custom", "_url": url},
            features=[],
        )

    # -- Worker ---------------------------------------------------------

    def _fetch_and_extract_one(self, url: str) -> tuple[str, WooProduct | None, str]:
        try:
            s, _, body = http_get_json(url, self.logger)
            if s != 200 or not isinstance(body, str):
                return url, None, f"status={s}"
            product = self._extract_product_from_html(url, body)
            return url, product, ""
        except Exception as e:
            return url, None, f"{type(e).__name__}: {e}"

    # -- Pipeline -------------------------------------------------------

    def build(self, max_products: int | None = None,
              concurrency: int = 2) -> list[WooProduct]:
        product_urls = self._get_product_urls()
        if not product_urls:
            raise HTTPError(
                f"Custom : aucune URL produit trouvée sur {self.base}. "
                f"Le site n'a peut-être pas de boutique en ligne, ou structure trop spécifique."
            )

        if max_products:
            product_urls = product_urls[:max_products]
            self.logger.info(f"  (limité à {len(product_urls)} pour ce run)")
        self.logger.info(f"  concurrence: {concurrency} thread(s)")

        from concurrent.futures import ThreadPoolExecutor, as_completed
        products: list[WooProduct] = []
        failures = 0
        done = 0
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {executor.submit(self._fetch_and_extract_one, url): url
                       for url in product_urls}
            for fut in as_completed(futures):
                done += 1
                url, product, err = fut.result()
                if product:
                    products.append(product)
                else:
                    failures += 1
                    if err:
                        self.logger.warning(f"[Custom] {url} {err}")
                if done % 10 == 0 or done == len(product_urls):
                    self.logger.info(f"  page {done}/{len(product_urls)} ({failures} échecs)")
        if failures:
            self.logger.warning(f"Custom terminé avec {failures}/{len(product_urls)} échecs")
        return products


# =============================================================================
# SECTION 6sexies — Scraper SumUp Store ({brand}.sumupstore.com)
# =============================================================================

class SumUpScraper:
    """Scrape une boutique SumUp Store (plateforme hosted, {brand}.sumupstore.com).

    SumUp utilise une structure HTML standardisée pour toutes les boutiques :
      - Sitemap index → sitemap.products.xml avec les URLs produits
      - URLs produits : /article/{slug}
      - Balises HTML avec attribut `data-selector="os-theme-product-X"` pour
        cibler proprement titre, prix, description, images, etc.
      - Open Graph présent en complément (og:title, og:description, og:image)

    Pas de variantes complexes pour la v1 (chaque article = 1 produit simple).
    """

    def __init__(self, brand_url: str, logger: logging.Logger):
        self.brand_url = brand_url
        p = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
        self.base = f"{p.scheme}://{p.netloc}"
        self.domain = p.netloc.replace("www.", "")
        self.logger = logger

    # -- URLs produits --------------------------------------------------

    def _get_product_urls(self) -> list[str]:
        try:
            s, _, body = http_get_json(f"{self.base}/sitemap.xml", self.logger)
        except HTTPError as e:
            raise HTTPError(f"SumUp : sitemap.xml inaccessible : {e}")
        if s != 200 or not isinstance(body, str):
            raise HTTPError(f"SumUp : sitemap.xml status={s}")
        sub_sitemaps = re.findall(r"<loc>([^<]+)</loc>", body)
        # Trouve sitemap.products.xml
        product_sm = next(
            (u for u in sub_sitemaps if "product" in u.lower()),
            None,
        )
        if not product_sm:
            raise HTTPError(
                "SumUp : pas de sub-sitemap 'products' trouvé dans sitemap.xml"
            )
        s2, _, body2 = http_get_json(product_sm, self.logger)
        if s2 != 200 or not isinstance(body2, str):
            raise HTTPError(f"SumUp : {product_sm} status={s2}")
        urls = re.findall(r"<loc>([^<]+)</loc>", body2)
        # Garde uniquement les /article/{slug}
        urls = [u for u in urls if "/article/" in u]
        self.logger.info(
            f"SumUp : sitemap utilisé ({product_sm}), {len(urls)} URLs produits"
        )
        return list(dict.fromkeys(urls))

    # -- Extraction page produit ---------------------------------------

    @staticmethod
    def _extract_by_selector(html_text: str, selector: str) -> str:
        """Extrait le contenu d'une balise avec data-selector="..."."""
        # On accepte l'attribut data-selector n'importe où dans la balise
        # et on prend le contenu jusqu'au prochain </tag>
        m = re.search(
            r'<(\w+)[^>]*\bdata-selector=["\']' + re.escape(selector)
            + r'["\'][^>]*>(.*?)</\1>',
            html_text, flags=re.DOTALL,
        )
        if not m:
            return ""
        content = m.group(2)
        # Strip nested HTML tags pour récupérer le texte
        content = re.sub(r"<[^>]+>", "", content)
        return html.unescape(content).strip()

    @staticmethod
    def _extract_all_by_selector(html_text: str, selector: str) -> list[str]:
        """Extrait toutes les valeurs avec un certain data-selector."""
        out = []
        for m in re.finditer(
            r'<(\w+)[^>]*\bdata-selector=["\']' + re.escape(selector)
            + r'["\'][^>]*>(.*?)</\1>',
            html_text, flags=re.DOTALL,
        ):
            text = re.sub(r"<[^>]+>", "", m.group(2))
            text = html.unescape(text).strip()
            if text:
                out.append(text)
        return out

    @staticmethod
    def _extract_og_meta(body: str) -> dict[str, str]:
        out = {}
        for prop in ("og:title", "og:description", "og:image"):
            m = re.search(
                r'<meta[^>]+(?:property|name)=["\']'
                + re.escape(prop) + r'["\'][^>]+content=["\']([^"\']+)["\']',
                body, flags=re.IGNORECASE,
            )
            if m:
                out[prop] = html.unescape(m.group(1)).strip()
        return out

    @staticmethod
    def _parse_price_str(price_str: str) -> tuple[float | None, str]:
        """Parse '12,80 €' → (12.80, 'EUR'). Gère les formats français/anglais."""
        if not price_str:
            return None, "EUR"
        # Currency
        currency = "EUR"
        if "€" in price_str or "EUR" in price_str.upper():
            currency = "EUR"
        elif "$" in price_str or "USD" in price_str.upper():
            currency = "USD"
        elif "£" in price_str or "GBP" in price_str.upper():
            currency = "GBP"
        # Numeric value
        m = re.search(r"(\d+(?:[.,]\d{1,2})?)", price_str.replace("\xa0", " "))
        if not m:
            return None, currency
        try:
            return float(m.group(1).replace(",", ".")), currency
        except ValueError:
            return None, currency

    @staticmethod
    def _extract_variants_from_rsc(body: str) -> dict | None:
        """Extrait le JSON variants depuis le React Server Components flight data.

        SumUp embed dans le HTML un JSON encodé en JS escape (\\") qui contient
        toutes les variantes avec leur prix, images, stock individuel. Format :

            \"variants\":{
              \"uuid1\":{\"uuid\":\"uuid1\",\"name\":\"BEIGE\",\"price\":1400,
                \"options\":[{\"name\":\"COULEURS\",\"value\":\"BEIGE\"}],
                \"images\":[\"https://...\"],\"quantity\":0,\"isAvailable\":false},
              \"uuid2\":{...},
              ...
            }

        Retourne le dict parsé {uuid: variant_data} ou None si non trouvé.
        """
        # Pattern : \"variants\":{
        m = re.search(r'\\"variants\\"\s*:\s*\{', body)
        if not m:
            return None
        start = m.end() - 1  # position du {

        # Balanced brace counting sur le JSON ESCAPED
        i = start
        depth = 0
        in_string = False
        while i < len(body):
            c = body[i]
            # \" = délimiteur de string dans le JSON encodé
            if c == '\\' and i + 1 < len(body):
                nxt = body[i + 1]
                if nxt == '"':
                    in_string = not in_string
                    i += 2
                    continue
                # Autre escape : skip 2 chars
                i += 2
                continue
            if not in_string:
                if c == '{':
                    depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0:
                        end = i + 1
                        raw = body[start:end]
                        # Décode les escapes JS : \\\" -> \" puis \" -> "
                        unescaped = raw.replace('\\\\', '\\').replace('\\"', '"')
                        try:
                            return json.loads(unescaped)
                        except json.JSONDecodeError:
                            return None
            i += 1
        return None

    @staticmethod
    def _extract_variant_groups(body: str) -> list[tuple[str, list[str]]]:
        """Extrait les groupes de variantes (selects SumUp).

        SumUp expose les axes de variation via :
          <select data-selector="os-theme-option-COULEURS">
            <option value="BEIGE">BEIGE</option>
            ...
          </select>

        Retourne : [(nom_groupe, [valeurs...]), ...]
        Ex: [("Couleurs", ["BEIGE", "BLANC", "BLEU", ...])]
        """
        groups: list[tuple[str, list[str]]] = []
        # Cherche tous les <select> avec data-selector="os-theme-option-..."
        for m in re.finditer(
            r'<select[^>]*data-selector=["\']os-theme-option-([^"\']+)["\'][^>]*>(.*?)</select>',
            body, flags=re.DOTALL | re.IGNORECASE,
        ):
            group_name_raw = m.group(1).strip()
            # Normalise le nom : "COULEURS" → "Couleurs"
            group_name = group_name_raw.replace("_", " ").replace("-", " ").strip()
            group_name = group_name.capitalize() if group_name.isupper() else group_name
            options_html = m.group(2)
            # Extrait les valeurs des options (skipper les options de placeholder vides)
            values = []
            for om in re.finditer(
                r'<option[^>]*value=["\']([^"\']*)["\'][^>]*>(.*?)</option>',
                options_html, flags=re.DOTALL,
            ):
                value = om.group(1).strip()
                label = re.sub(r"<[^>]+>", "", om.group(2)).strip()
                # On utilise le label affiché (visible utilisateur) plutôt que value
                # car SumUp met parfois la même valeur dans les deux
                chosen = label or value
                if chosen and chosen not in values:
                    values.append(chosen)
            if values:
                groups.append((group_name, values))
        return groups

    def _extract_product_from_html(self, url: str, body: str) -> WooProduct | None:
        # Nom : data-selector="os-theme-product-title" en priorité, sinon og:title
        title = self._extract_by_selector(body, "os-theme-product-title")
        og = self._extract_og_meta(body)
        if not title:
            title = og.get("og:title") or ""

        if not title:
            self.logger.warning(f"[SumUp] {url} : ni titre os-theme ni og:title trouvés")
            return None

        # Prix : data-selector="os-theme-product-price"
        price_str = self._extract_by_selector(body, "os-theme-product-price")
        price, currency = self._parse_price_str(price_str)
        parent_price_cents = str(int(round(price * 100))) if price else None

        # Description : essaie plusieurs selectors SumUp + og:description
        description = ""
        for sel in (
            "os-theme-product-description",
            "os-theme-product-short-description",
            "os-theme-article-description",
        ):
            v = self._extract_by_selector(body, sel)
            if v and len(v) > len(description):
                description = v
        if not description:
            description = og.get("og:description") or ""

        # Images : data-selector os-theme-product-image (toutes), + og:image
        images_urls = []
        for sel in (
            "os-theme-product-image",
            "os-theme-product-gallery-image",
            "os-theme-product-cover",
        ):
            for m in re.finditer(
                r'<(?:img|a)[^>]*\bdata-selector=["\']' + re.escape(sel)
                + r'["\'][^>]*(?:src|href)=["\']([^"\']+)["\']',
                body, flags=re.IGNORECASE,
            ):
                u = m.group(1)
                if u and u not in images_urls:
                    images_urls.append(u)
        # Fallback : images <img> avec src vers images.sumup.com
        for m in re.finditer(
            r'<img[^>]+src=["\'](https?://images\.sumup\.com/[^"\']+)["\']',
            body,
        ):
            u = m.group(1)
            if u not in images_urls:
                images_urls.append(u)
        if og.get("og:image") and og["og:image"] not in images_urls:
            images_urls.append(og["og:image"])
        images = [{"src": u} for u in images_urls]

        # Slug depuis l'URL (= SKU fallback)
        slug_match = re.search(r"/article/([^/?#]+)", url)
        slug = slug_match.group(1) if slug_match else ""
        product_id = abs(hash(slug)) % (10 ** 9)
        sku = slug  # SumUp ne fournit pas de SKU séparé

        # Stock : pas d'info exposée → assume in_stock
        in_stock = True

        # VARIANTES : 2 sources possibles, par ordre de richesse :
        #
        # 1. JSON RSC inline (PRIORITAIRE) : SumUp embed un JSON complet avec
        #    prix/images/stock individuels par variante. C'est la source idéale.
        # 2. Selects HTML : fallback simple, on prend juste les valeurs et on
        #    duplique le prix parent.
        variations_data: list[dict] = []
        rsc_variants = self._extract_variants_from_rsc(body)
        if rsc_variants and isinstance(rsc_variants, dict) and len(rsc_variants) > 1:
            self.logger.debug(
                f"[SumUp] {url} : JSON RSC trouvé avec {len(rsc_variants)} variantes"
            )
            for idx, (uuid, v) in enumerate(rsc_variants.items()):
                if not isinstance(v, dict):
                    continue
                v_name = (v.get("name") or "").strip()
                v_price_cents_int = v.get("price") or 0
                # Si prix variante = 0, fallback sur le prix parent
                v_price_cents = (
                    str(v_price_cents_int) if v_price_cents_int else parent_price_cents
                )
                # Attributs : v.options = [{"name": "COULEURS", "value": "BEIGE"}, ...]
                attrs = []
                for opt in (v.get("options") or []):
                    if isinstance(opt, dict) and opt.get("name") and opt.get("value"):
                        # Normalise le nom de groupe : "COULEURS" → "Couleurs"
                        gname = opt["name"]
                        if gname.isupper():
                            gname = gname.capitalize()
                        attrs.append({"name": gname, "value": opt["value"].strip()})
                # Images variante (peut être vide si la marque n'a pas mis d'image)
                v_imgs = [{"src": u} for u in (v.get("images") or []) if u]
                # Stock
                qty = v.get("quantity")
                is_avail = bool(v.get("isAvailable", True))
                # SKU : use UUID ou construit
                if attrs:
                    sku_suffix = "-".join(
                        re.sub(r"[^a-zA-Z0-9]", "", a["value"])[:10] for a in attrs
                    )
                    v_sku = f"{slug}-{sku_suffix}"
                else:
                    v_sku = uuid[:50]
                variations_data.append({
                    "id": idx,
                    "sku": v_sku,
                    "prices": {
                        "price": v_price_cents,
                        "regular_price": v_price_cents,
                        "currency_code": currency,
                        "currency_minor_unit": 2,
                    },
                    "is_in_stock": is_avail,
                    "low_stock_remaining": qty if isinstance(qty, int) and qty >= 0 else None,
                    "images": v_imgs,
                    "_parent_attributes": attrs,
                })
        else:
            # Fallback : extraction des selects (sans prix/images par variante)
            groups = self._extract_variant_groups(body)
            if groups:
                from itertools import product as iproduct
                group_names = [g[0] for g in groups]
                group_values = [g[1] for g in groups]
                combos = list(iproduct(*group_values))
                self.logger.debug(
                    f"[SumUp] {url} : pas de JSON RSC, fallback selects "
                    f"({len(groups)} axes × {[len(v) for v in group_values]} → "
                    f"{len(combos)} variantes)"
                )
                for idx, combo in enumerate(combos):
                    attrs = []
                    sku_parts = []
                    for gname, gval in zip(group_names, combo):
                        attrs.append({"name": gname, "value": gval})
                        sku_parts.append(re.sub(r"[^a-zA-Z0-9]", "", gval)[:10])
                    v_sku = f"{slug}-{'-'.join(sku_parts)}" if sku_parts else slug
                    variations_data.append({
                        "id": idx,
                        "sku": v_sku,
                        "prices": {
                            "price": parent_price_cents,
                            "regular_price": parent_price_cents,
                            "currency_code": currency,
                            "currency_minor_unit": 2,
                        },
                        "is_in_stock": True,
                        "low_stock_remaining": None,
                        "images": [],
                        "_parent_attributes": attrs,
                    })

        p_type = "variable" if len(variations_data) > 1 else "simple"
        variations_ids = [v["id"] for v in variations_data]

        return WooProduct(
            id=product_id,
            name=title,
            slug=slug,
            type=p_type,
            sku=sku,
            short_description=description,
            description="",
            permalink=url,
            prices={
                "price": parent_price_cents,
                "currency_code": currency,
                "currency_minor_unit": 2,
            } if parent_price_cents else {},
            images=images,
            categories=[],
            tags=[],
            attributes=[],
            variations_ids=variations_ids,
            variations_data=variations_data,
            is_in_stock=in_stock,
            low_stock_remaining=None,
            raw={"_source": "sumup", "_url": url},
            features=[],
        )

    def _fetch_and_extract_one(self, url: str) -> tuple[str, WooProduct | None, str]:
        try:
            s, _, body = http_get_json(url, self.logger)
            if s != 200 or not isinstance(body, str):
                return url, None, f"status={s}"
            product = self._extract_product_from_html(url, body)
            return url, product, ""
        except Exception as e:
            return url, None, f"{type(e).__name__}: {e}"

    def build(self, max_products: int | None = None,
              concurrency: int = 2) -> list[WooProduct]:
        product_urls = self._get_product_urls()
        if not product_urls:
            raise HTTPError(f"SumUp : aucune URL produit sur {self.base}")
        if max_products:
            product_urls = product_urls[:max_products]
            self.logger.info(f"  (limité à {len(product_urls)} pour ce run)")
        self.logger.info(f"  concurrence: {concurrency} thread(s)")

        from concurrent.futures import ThreadPoolExecutor, as_completed
        products: list[WooProduct] = []
        failures = 0
        done = 0
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {executor.submit(self._fetch_and_extract_one, url): url
                       for url in product_urls}
            for fut in as_completed(futures):
                done += 1
                url, product, err = fut.result()
                if product:
                    products.append(product)
                else:
                    failures += 1
                    if err:
                        self.logger.warning(f"[SumUp] {url} {err}")
                if done % 10 == 0 or done == len(product_urls):
                    self.logger.info(f"  page {done}/{len(product_urls)} ({failures} échecs)")
        if failures:
            self.logger.warning(f"SumUp terminé avec {failures}/{len(product_urls)} échecs")
        return products


# =============================================================================
# SECTION 7 — Nettoyage des descriptions
# =============================================================================

# Pattern shortcodes Divi (`[et_pb_section ...]`, `[/et_pb_text]`...)
DIVI_SHORTCODE = re.compile(r"\[\/?et_pb_[a-z0-9_]+[^\]]*\]", re.IGNORECASE)
# Autres shortcodes WP générique
WP_SHORTCODE = re.compile(r"\[\/?[a-z][a-z0-9_-]*(?:\s+[^\]]*)?\]", re.IGNORECASE)
# Tags HTML
HTML_TAG = re.compile(r"<[^>]+>")
# Multi-whitespace
WHITESPACE = re.compile(r"\s+")
# Caractères de contrôle / NBSP / weird Divi-encoded quotes
WEIRD_CHARS = re.compile(r"[ ​‌‍﻿]+")


def _dedupe_products_by_name(products: list[WooProduct], logger: logging.Logger) -> list[WooProduct]:
    """Dédoublonne les produits ayant le même nom normalisé.

    Beaucoup de marques créent par erreur des doublons dans leur back-office
    (ex: "Monoi passion" et "Monoi Passion" — juste différence de casse).
    On garde la 1ère occurrence et on drop les suivantes.

    Normalisation : lowercase + sans accents + espaces multiples écrasés en un seul.
    """
    if len(products) < 2:
        return products

    seen: dict[str, WooProduct] = {}
    dropped: list[str] = []
    for p in products:
        key = _normalize(p.name).strip()
        # Écrase espaces multiples
        key = re.sub(r"\s+", " ", key)
        if not key:
            # Sans nom → garde tel quel (cas rare)
            seen[f"_noname_{id(p)}"] = p
            continue
        if key in seen:
            dropped.append(p.name)
        else:
            seen[key] = p

    if dropped:
        logger.info(
            f"Dédoublonnage : {len(dropped)} produit(s) avec nom déjà existant retiré(s). "
            f"Exemples : {dropped[:3]}"
        )
    return list(seen.values())


def _strip_common_brand_suffix(products: list[WooProduct], logger: logging.Logger) -> None:
    """Détecte et retire un (ou plusieurs) suffixe(s) commun(s) dans les noms.

    Cas typiques :
      - " • Le comptoir des saveurs à Sarrebourg" (suffixe site automatique)
      - " à base de Rhum" (suffixe SEO mal configuré sur une partie du catalogue)

    Stratégie :
      Pour chaque séparateur ( • / | / - / – / : / · / — ), on regarde les
      suffixes qui apparaissent dans au moins max(5, 30%) des noms. On les
      retire UNIQUEMENT des noms qui les contiennent. On répète jusqu'à
      stabilisation (max 3 passes pour éviter une boucle).

    Modifie les WooProduct en place.
    """
    if len(products) < 5:
        return

    SEPARATORS = (" • ", " | ", " - ", " – ", " · ", " : ", " — ")
    MIN_OCCURRENCES_ABS = 5
    MIN_OCCURRENCES_PCT = 0.30

    for pass_num in range(3):  # max 3 passes pour catcher les suffixes en cascade
        names = [p.name for p in products if p.name]
        if len(names) < 5:
            return
        threshold = max(MIN_OCCURRENCES_ABS, int(MIN_OCCURRENCES_PCT * len(names)))
        candidates: dict[str, int] = {}

        for sep in SEPARATORS:
            suffixes = []
            for n in names:
                idx = n.rfind(sep)
                if idx > 0:
                    s = sep + n[idx + len(sep):]
                    suffixes.append(s)
            if not suffixes:
                continue
            from collections import Counter
            for suffix, count in Counter(suffixes).items():
                if count >= threshold:
                    # Garde le plus long s'il existe plusieurs candidats qui se chevauchent
                    candidates[suffix] = count

        if not candidates:
            return

        # On prend tous les candidats triés par longueur décroissante
        # pour retirer les plus longs en premier
        sorted_candidates = sorted(candidates.items(), key=lambda x: -len(x[0]))
        any_change = False
        for suffix, count in sorted_candidates:
            # Re-check live count après les modifs précédentes du même passage
            current_matches = [p for p in products if p.name.endswith(suffix)]
            if len(current_matches) < threshold:
                continue
            logger.info(
                f"Suffixe commun retiré ({len(current_matches)} produits, "
                f"passe {pass_num + 1}) : {suffix!r}"
            )
            for p in current_matches:
                p.name = p.name[: -len(suffix)].rstrip()
                any_change = True
        if not any_change:
            return


# Patterns d'artefacts ChatGPT/Claude/Tailwind qui ont fui via du copier-coller
# dans le back-office des marques (cas vu sur Cosmella). Si une ligne contient
# un de ces patterns, c'est du résidu de markup et on la jette.
JUNK_LINE_PATTERNS = [
    re.compile(r'\b(?:data-turn-id|data-testid|data-scroll-anchor|data-turn|data-start|data-end|data-is-last-node|data-is-only-node|data-message-author-role|data-message-id|data-writing-block)\b'),
    re.compile(r'content-visibility'),
    re.compile(r'pointer-events-(?:auto|none)'),
    re.compile(r'conversation-turn'),
    re.compile(r'thread-(?:response|scroll-vars|content-margin|content-max-width)'),
    re.compile(r'scroll-(?:mb|mt|root)-\['),
    re.compile(r'var\(--(?:thread|scroll|header|spacing|shadow)-'),
    re.compile(r'\[--thread-'),
    re.compile(r'calc\(var\(--'),
    re.compile(r'\bagent-turn\b'),
    re.compile(r'request-[0-9a-f]{8}-[0-9a-f]{4}-'),
    re.compile(r'\[contain-intrinsic-size'),
    re.compile(r'token-text-primary'),
    re.compile(r'has-data-writing-block'),
    re.compile(r'\bclass\s*=\s*["\'][^"\']*(?:pointer|whitespace|scroll|outline|focus|gap-|max-w-|min-h-|@w-|prose-)'),
    # Squarespace : blocs CSS inline (#block-{hash}, --tweak-*, @media, .sqs-html-content)
    re.compile(r'#block-[a-f0-9]{12,}'),
    re.compile(r'--tweak-[a-z-]+'),
    re.compile(r'\.sqs-html-content'),
    re.compile(r'@media\s+screen'),
    re.compile(r'mix-blend-mode\s*:'),
    re.compile(r'box-sizing\s*:'),
    re.compile(r'border-radius\s*:\s*var\('),
]


def _is_junk_line(line: str) -> bool:
    """Détecte une ligne qui est en fait du markup résiduel (CSS Tailwind, attributs HTML ChatGPT)."""
    stripped = line.strip()
    if not stripped:
        return False
    for p in JUNK_LINE_PATTERNS:
        if p.search(stripped):
            return True
    # Heuristique : ligne dominée par des caractères de markup (crochets, parens, deux-points)
    # sans assez de lettres pour être de la prose.
    if len(stripped) > 40:
        special = sum(1 for c in stripped if c in "[](){}:;=_")
        letters = sum(1 for c in stripped if c.isalpha())
        if letters > 0 and special / letters > 0.5:
            return True
    return False


def _clean_text(s: str) -> str:
    """Nettoie un blob HTML/Divi/entités en texte propre.

    ORDRE CRITIQUE :
      0. Strip <style>...</style> et <script>...</script> ENTIÈREMENT (contenu inclus)
      1. Marqueurs de paragraphe (<br>, </p>, </li>, </div>, </section>) -> \n
      2. Strip TAGS HTML AVANT unescape (sinon &gt; dans les attributs casse le regex)
      3. Strip shortcodes Divi/WP
      4. Décode les entités HTML
      5. Normalise espaces zero-width / NBSP
      6. Drop les lignes "junk" (CSS Tailwind, attributs ChatGPT, blocks Squarespace etc.)
    """
    if not s:
        return ""

    # 0. Strip <style> et <script> blocks entiers (contenu compris) AVANT toute autre
    # transformation. Sinon le contenu CSS/JS reste comme du texte dans le résultat.
    # Cas typique sur Squarespace : <style>#block-xxx { --tweak-* ... }</style> au
    # milieu du body produit.
    s = re.sub(r"<style\b[^>]*>.*?</style>", " ", s, flags=re.DOTALL | re.IGNORECASE)
    s = re.sub(r"<script\b[^>]*>.*?</script>", " ", s, flags=re.DOTALL | re.IGNORECASE)

    # 1. Markers de paragraphe en \n (avant strip général)
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.IGNORECASE)
    s = re.sub(r"</(?:p|li|div|section|article|header|footer|h[1-6])\s*>", "\n",
               s, flags=re.IGNORECASE)

    # 2. Strip HTML tags AVANT unescape (sinon &gt; encodé devient > et perturbe la regex)
    s = HTML_TAG.sub("", s)

    # 3. Strip shortcodes
    s = DIVI_SHORTCODE.sub(" ", s)
    s = WP_SHORTCODE.sub(" ", s)

    # 4. Décode les entités HTML
    s = html.unescape(s)

    # 5. Normalise espaces zero-width / NBSP / spéciaux
    s = WEIRD_CHARS.sub(" ", s)

    # 6. Nettoie ligne par ligne
    kept = []
    for ln in s.splitlines():
        ln = ln.strip()
        if not ln:
            continue
        if _is_junk_line(ln):
            continue
        ln = re.sub(r"[ \t]+", " ", ln)  # normalise espaces internes
        kept.append(ln)

    return "\n".join(kept).strip()


# Patterns de headers de section pour l'extraction INCI / ingrédients
INCI_HEADER_RE = re.compile(
    r'^\s*(?:Ingr[éeè]dients?\s*INCI|Liste\s+INCI|INCI|Composition\s+INCI|Ingr[éeè]dients?|Composition|Ingredients\s+list|Ingredients)\s*[:\.\-]?\s*$',
    re.IGNORECASE | re.MULTILINE,
)


def extract_inci(text: str) -> tuple[str, str]:
    """Cherche une section 'Ingrédients/INCI/Composition' dans le texte.

    Retourne (text_sans_inci, contenu_inci). Si rien trouvé, retourne (text, "").
    Détecte un header type 'Ingrédients :' ou 'INCI' sur sa propre ligne, puis
    capture les lignes suivantes jusqu'à la prochaine ligne vide ou un autre
    header en titre.
    """
    if not text:
        return "", ""

    m = INCI_HEADER_RE.search(text)
    if not m:
        return text, ""

    start = m.end()
    # On regarde les lignes après le header. On capture jusqu'à :
    #   - une ligne vide
    #   - un nouveau header (ligne courte qui se termine par : ou .)
    remainder = text[start:]
    lines = remainder.split("\n")
    captured = []
    for ln in lines:
        ln_stripped = ln.strip()
        if not ln_stripped:
            break  # ligne vide = fin de section
        # Nouveau header probable ? (ligne courte avec : en fin)
        if len(ln_stripped) < 60 and ln_stripped.endswith(":"):
            break
        captured.append(ln_stripped)
        if len(captured) >= 30:  # safety cap
            break

    inci_content = " ".join(captured).strip()
    if len(inci_content) < 5:
        return text, ""  # trop court pour être un vrai INCI

    # On retire le header + le contenu capturé du texte
    end_offset = start + sum(len(l) + 1 for l in lines[:len(captured)])
    cleaned = (text[:m.start()] + text[end_offset:]).strip()
    # Nettoie les doubles \n laissés
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned, inci_content


def clean_description(short: str, full: str) -> str:
    """Construit la description Ankorstore en concaténant short + full nettoyés.

    Règle (suite feedback Célia) :
      - Nettoie les deux indépendamment
      - Si l'un contient l'autre, on garde le plus long (évite la duplication)
      - Sinon on concatène : short (résumé) + \n\n + full (détails)
    """
    cleaned_short = _clean_text(short)
    cleaned_full = _clean_text(full)

    if not cleaned_short and not cleaned_full:
        return ""
    if not cleaned_short:
        return cleaned_full
    if not cleaned_full:
        return cleaned_short

    # Dédoublonnage
    if cleaned_short in cleaned_full:
        return cleaned_full
    if cleaned_full in cleaned_short:
        return cleaned_short

    # Sinon concaténation
    return cleaned_short + "\n\n" + cleaned_full


# =============================================================================
# SECTION 8 — Classification des attributs (taille / couleur / autre)
# =============================================================================

def _normalize(s: str) -> str:
    """Lowercase + remove accents-like chars pour matching robuste."""
    if not s:
        return ""
    s = s.lower()
    # Remove diacritics simple
    for a, b in (("à", "a"), ("â", "a"), ("ä", "a"), ("é", "e"), ("è", "e"), ("ê", "e"),
                 ("ë", "e"), ("î", "i"), ("ï", "i"), ("ô", "o"), ("ö", "o"), ("ù", "u"),
                 ("û", "u"), ("ü", "u"), ("ç", "c"), ("ñ", "n")):
        s = s.replace(a, b)
    return s


def classify_attribute(name: str) -> str:
    """Retourne 'size', 'color' ou 'other'."""
    n = _normalize(name)
    if any(kw in n for kw in (_normalize(k) for k in SIZE_KEYWORDS)):
        return "size"
    if any(kw in n for kw in (_normalize(k) for k in COLOR_KEYWORDS)):
        return "color"
    return "other"


# =============================================================================
# SECTION 9 — Filtre non-marchandises
# =============================================================================

def is_non_merchandise(p: WooProduct) -> tuple[bool, str]:
    """Retourne (True/False, raison). On veut filtrer ateliers/services/etc."""
    name_n = _normalize(p.name)
    for kw in NON_MERCHANDISE_KEYWORDS_TITLE:
        if _normalize(kw) in name_n:
            return True, f"titre contient '{kw}'"
    for cat in p.categories:
        cat_n = _normalize(cat.get("name") or "")
        for kw in NON_MERCHANDISE_KEYWORDS_CATEGORY:
            if _normalize(kw) in cat_n:
                return True, f"catégorie '{cat.get('name')}' contient '{kw}'"
    return False, ""


# =============================================================================
# SECTION 10 — Génération de SKU
# =============================================================================

def make_sku(parent_slug: str, parent_sku: str, variation_id: int | None,
             variation_index: int) -> str:
    """Génère un SKU stable et unique. Max 50 chars."""
    # Si le parent a un SKU et qu'on a une variation, on combine
    if parent_sku:
        base = parent_sku
        if variation_id is not None:
            base = f"{parent_sku}-{variation_id}"
    else:
        # Slugifie le slug existant + variation_id
        base = parent_slug or "product"
        if variation_id is not None:
            base = f"{base}-{variation_id}"
        elif variation_index is not None:
            base = f"{base}-{variation_index}"
    # Cap à SKU_MAX_LEN
    if len(base) > SKU_MAX_LEN:
        # On garde la fin (l'ID variation est plus discriminant que le début du slug)
        # Trim depuis le début du slug
        suffix_part = ""
        for sep in ("-",):
            idx = base.rfind(sep)
            if 0 < idx < len(base):
                suffix_part = base[idx:]
                break
        if suffix_part and len(suffix_part) < SKU_MAX_LEN:
            keep = SKU_MAX_LEN - len(suffix_part)
            base = base[:keep] + suffix_part
        else:
            base = base[-SKU_MAX_LEN:]
    return base


# =============================================================================
# SECTION 11 — Détection des tags booléens
# =============================================================================

def _keyword_matches(haystack_n: str, keyword: str) -> bool:
    """Match avec word boundaries pour éviter les faux positifs.

    Exemple : "ipa" ne matchera PAS dans "frangipani" car ce n'est pas un mot
    isolé. "porter" ne matchera PAS dans "emporter".
    """
    kw_n = _normalize(keyword)
    pattern = r"\b" + re.escape(kw_n) + r"\b"
    return bool(re.search(pattern, haystack_n))


def detect_booleans(p: WooProduct, cleaned_desc: str) -> dict[str, bool]:
    """Pour chaque tag, vérifie deux scopes :

    - `anywhere` : nom + description + catégories + tags
    - `name_or_category_only` : nom + catégories uniquement (évite faux positifs
       sur des sous-chaînes communes type 'emporter' qui contient 'porter')

    Tous les matchs utilisent des word boundaries — pas de substring matching.
    Un tag est activé si au moins UN keyword anywhere matche OU au moins UN
    keyword name_or_category matche, ET aucun keyword négatif matche.
    """
    cats_text = " ".join(c.get("name") or "" for c in p.categories)
    tags_text = " ".join(t.get("name") or "" for t in p.tags)
    haystack_full = " ".join([p.name, cleaned_desc, cats_text, tags_text])
    haystack_name_cat = " ".join([p.name, cats_text])

    n_full = _normalize(haystack_full)
    n_name_cat = _normalize(haystack_name_cat)

    out: dict[str, bool] = {}
    for col, anywhere, name_cat_only, negatives in BOOLEAN_TAGS:
        # Si un négatif est présent (word boundary), on ne coche pas
        if any(_keyword_matches(n_full, kw) for kw in negatives):
            out[col] = False
            continue
        # Match anywhere (word boundary) OU match name/category (word boundary)
        out[col] = (
            any(_keyword_matches(n_full, kw) for kw in anywhere)
            or any(_keyword_matches(n_name_cat, kw) for kw in name_cat_only)
        )
    return out


# =============================================================================
# SECTION 12 — Mapping vers la matrice Ankorstore
# =============================================================================

@dataclass
class AnkorRow:
    """Une ligne de la matrice Ankorstore (1 ligne = 1 variante)."""
    data: dict[str, Any] = field(default_factory=dict)

    def set(self, col: str, val: Any) -> None:
        self.data[col] = val

    def get(self, col: str) -> Any:
        return self.data.get(col)


def cents_to_decimal(price_str: str | int | None, minor_unit: int = 2) -> float | None:
    """Convertit '1500' (centimes) en 15.00."""
    if price_str is None or price_str == "":
        return None
    try:
        cents = int(str(price_str))
    except (ValueError, TypeError):
        try:
            return float(price_str)
        except (ValueError, TypeError):
            return None
    return cents / (10 ** minor_unit)


def _clean_feature_value(value: str) -> str:
    """Nettoie une valeur de feature : strip HTML, décode entités, normalise."""
    if not value:
        return ""
    # Convertit <br> / </br> en saut de ligne avant strip
    v = re.sub(r"</?br\s*/?>", "\n", value, flags=re.IGNORECASE)
    # Strip autres tags HTML
    v = re.sub(r"<[^>]+>", "", v)
    # Décode entités
    v = html.unescape(v)
    # Normalise espaces
    v = re.sub(r"[ \t]+", " ", v)
    # Conserve les sauts de ligne mais nettoie autour
    lines = [ln.strip() for ln in v.splitlines() if ln.strip()]
    return "\n".join(lines)


def upscale_image_url(url: str) -> str:
    """Transforme une URL d'image thumbnail vers la version haute résolution.

    Cas gérés :
      - WooCommerce : `image-300x300.jpg` → `image.jpg` (image source originale)
      - PrestaShop  : `123-home_default/p.jpg` → `123/p.jpg` (image originale,
        souvent 2400×2400+) au lieu de `large_default` (~458×458, sous le min
        Ankorstore 500×500).

    Note : Ankorstore exige min 500×500 px. Retirer le suffix donne l'image
    originale, qui est presque toujours largement au-dessus du minimum.
    """
    if not url or not isinstance(url, str):
        return url
    # WooCommerce / WordPress : retire le suffix dimensionnel `-WxH` juste avant l'extension
    url = re.sub(r'(-\d+x\d+)(\.\w+)(\?.*)?$', r'\2\3', url)
    # PrestaShop : SUPPRIME le suffix `-XXX_default` complètement (donne l'image
    # originale haute résolution).
    url = re.sub(
        r'(/\d+(?:-\d+)?)-(?:small|medium|home|cart|miniature|large|thickbox|big)_default/',
        r'\1/',
        url,
    )
    # Wix (static.wixstatic.com) : remplace w_XXX,h_YYY par w_2000,h_2000
    # pour avoir une résolution suffisante pour Ankorstore (min 500×500).
    if "wixstatic.com" in url:
        url = re.sub(r"/v1/(fit|fill|crop)/w_\d+,h_\d+",
                     r"/v1/\1/w_2000,h_2000", url)
    return url


def route_features(features: list[dict]) -> dict[str, Any]:
    """Route les features structurées (PrestaShop) vers les colonnes Ankorstore.

    Retourne un dict {col_ankorstore: valeur} avec les colonnes :
      Dimensions, Composition, Matériau, Poids, Volume, Fabriqué en

    Plus une clé spéciale "_extra_text" : string à appender à la description
    (pour les features non mappables : Lavage, Garnissage, OEKO-TEX, etc.).
    """
    out: dict[str, Any] = {"_extra_text": ""}
    extras: list[str] = []

    # Mapping nom_feature_normalisé -> col_ankorstore (match exact ou substring)
    name_to_col = {
        "matiere": "Composition",
        "matiere principale": "Composition",
        "matiere secondaire": "Composition",
        "composition": "Composition",
        "garnissage": "Composition",
        "tissu": "Composition",
        "materiau": "Matériau",
        "materiaux": "Matériau",
        "material": "Matériau",
        "volume": "Volume",
        "contenance": "Volume",
        "capacite": "Volume",
        "couleur": "Couleurs des variants",
        "color": "Couleurs des variants",
        "colour": "Couleurs des variants",
        "coloris": "Couleurs des variants",
        "senteur": "Autres attributs de variante",
        "parfum": "Autres attributs de variante",
        "fragrance": "Autres attributs de variante",
        "scent": "Autres attributs de variante",
    }

    # Features qui vont en "Fabriqué en" (avec conversion pays → ISO)
    country_feature_names = {
        "lieu de fabrication", "pays d'origine", "pays de fabrication",
        "origine", "fabrique en", "fabriqué en", "made in", "country of origin",
    }

    # Features qu'on append toujours à la description (entretien, etc.)
    append_to_desc_names = ("lavage", "entretien")

    # Dimensions composites : on accumule hauteur/largeur/profondeur séparées
    # puis on les combine en un seul string final.
    dim_parts: list[tuple[str, str]] = []  # [("H", "8,3 cm"), ("L", "6,8 cm"), ...]
    dim_keys_seen: set[str] = set()

    def _extract_unit_from_name(label: str) -> str:
        """Extrait l'unité du nom de feature : 'Hauteur (en cm)' → ' cm'."""
        m = re.search(r"\(en\s+([a-z]{1,4})\)", label, re.IGNORECASE)
        if m:
            return f" {m.group(1).lower()}"
        return ""

    composition_parts: list[str] = []
    weight_value: str | None = None

    for f in features:
        if not isinstance(f, dict):
            continue
        name = (f.get("name") or "").strip()
        value = _clean_feature_value(f.get("value") or "")
        if not name or not value:
            continue
        name_norm = _normalize(name)

        # Pays d'origine
        if name_norm in country_feature_names or any(c in name_norm for c in country_feature_names):
            iso = COUNTRY_TO_ISO.get(_normalize(value))
            if iso:
                out["Fabriqué en"] = iso
            else:
                if len(value) <= 3 and value.isalpha():
                    out["Fabriqué en"] = value.upper()
                else:
                    extras.append(f"{name}: {value}")
            continue

        # Dimensions composites : Hauteur / Largeur / Profondeur / Longueur / Diamètre
        dim_match = None
        for dim_kw, dim_label in (
            ("hauteur", "H"), ("largeur", "L"), ("profondeur", "P"),
            ("longueur", "Lng"), ("diametre", "Ø"), ("diameter", "Ø"),
            ("epaisseur", "Ép"),
        ):
            if dim_kw in name_norm and dim_kw not in dim_keys_seen:
                unit = _extract_unit_from_name(name)
                dim_parts.append((dim_label, f"{value}{unit}"))
                dim_keys_seen.add(dim_kw)
                dim_match = True
                break
        if dim_match:
            continue

        # Dimensions directes (la feature s'appelle "Dimensions" elle-même)
        if name_norm in ("dimensions", "dimension", "taille") and "Dimensions" not in out:
            out["Dimensions"] = value
            continue

        # Poids
        if "poids" in name_norm or "weight" in name_norm:
            if weight_value is None:
                unit = _extract_unit_from_name(name)
                weight_value = f"{value}{unit}".strip()
            continue

        # Mapping direct nom→colonne
        mapped_col = name_to_col.get(name_norm)
        if not mapped_col:
            # Match partiel (substring) pour matiere, couleur, etc.
            for kw, col in name_to_col.items():
                if kw in name_norm:
                    mapped_col = col
                    break
        if mapped_col:
            if mapped_col == "Composition":
                composition_parts.append(f"{name}: {value}")
            else:
                # Pour les colonnes uniques, on garde la 1ère valeur trouvée
                if mapped_col not in out:
                    out[mapped_col] = value
            continue

        # Entretien / lavage / etc. → append à description
        if any(kw in name_norm for kw in append_to_desc_names):
            extras.append(f"{name}: {value}")
            continue

        # Autres features : append à description
        extras.append(f"{name}: {value}")

    # Compile dimensions composites
    if dim_parts:
        out["Dimensions"] = " × ".join(f"{lbl} {v}" for lbl, v in dim_parts)

    if weight_value is not None:
        out["Poids"] = weight_value

    if composition_parts:
        out["Composition"] = " | ".join(composition_parts)

    if extras:
        out["_extra_text"] = "\n".join(extras)

    return out


def stock_value(in_stock: bool, low_stock: int | None) -> int | str:
    """Retourne le stock à mettre dans la matrice.

    Convention Ankorstore :
      - 0 si explicitement out_of_stock
      - quantité exacte si connue
      - "in_stock" (string) si en stock sans quantité précise
    """
    if not in_stock:
        return 0
    if isinstance(low_stock, int) and low_stock >= 0:
        return low_stock
    return "in_stock"


MAX_IMAGES_PER_PRODUCT = 15  # cap dur (template Ankorstore limite à 15 colonnes images)


def extract_image_urls(images: list[dict], limit: int = MAX_IMAGES_PER_PRODUCT) -> list[str]:
    """Extrait les URLs images. Applique l'upscaling pour atteindre une résolution
    suffisante (Ankorstore exige min 500×500). Cap par défaut à 15 (max template)."""
    out: list[str] = []
    for img in images[:limit]:
        if isinstance(img, dict):
            src = img.get("src") or img.get("thumbnail") or ""
            if src:
                out.append(upscale_image_url(src))
    return out


def build_rows_for_product(
    p: WooProduct,
    logger: logging.Logger,
) -> list[AnkorRow]:
    """Transforme un WooProduct en N lignes AnkorRow (1 par variante).

    Cas :
      - simple    -> 1 ligne (variation_id=None, attributs vides)
      - variable  -> N lignes (1 par variation_data)
      - autre     -> 1 ligne (best effort)
    """
    rows: list[AnkorRow] = []

    # Champs communs
    cleaned_desc = clean_description(p.short_description, p.description)
    # Si une section INCI/Ingrédients est détectée, on l'extrait dans une variable
    # séparée pour la mettre dans la colonne "Liste INCI" du template
    cleaned_desc, inci_content = extract_inci(cleaned_desc)
    if inci_content:
        logger.debug(f"#{p.id} '{p.name}': section INCI/Ingrédients extraite "
                     f"({len(inci_content)} chars)")

    # Routing des features structurées (PrestaShop principalement) vers les
    # colonnes Ankorstore : Dimensions, Composition, Matériau, Poids, Volume,
    # Fabriqué en. Le reste (entretien, etc.) est appendé à la description.
    feature_cols = route_features(p.features)
    extra_desc_text = feature_cols.pop("_extra_text", "")
    if extra_desc_text:
        cleaned_desc = (cleaned_desc + "\n\n" + extra_desc_text).strip()

    if len(cleaned_desc) < 30:
        logger.warning(f"#{p.id} '{p.name}': description nettoyée < 30 chars "
                       f"(len={len(cleaned_desc)}). Sera signalée comme à compléter.")
    parent_images = extract_image_urls(p.images, limit=MAX_IMAGES_PER_PRODUCT)
    if not parent_images:
        logger.warning(f"#{p.id} '{p.name}': aucune image — Image 1 obligatoire manquante")

    booleans = detect_booleans(p, cleaned_desc)
    name = html.unescape(p.name).strip()

    # Détermine si on a des variations exploitables
    var_count = len(p.variations_data)

    if p.type == "variable" and var_count > 0:
        # Détecte si les SKU des variantes sont uniques entre elles.
        # Beaucoup de marques mettent le même SKU sur toutes les variantes (= SKU parent
        # répliqué). Dans ce cas il faut générer des SKU uniques pour Ankorstore.
        raw_v_skus = [(vd.get("sku") or "").strip() for vd in p.variations_data]
        non_empty_skus = [s for s in raw_v_skus if s]
        # Conditions pour utiliser les SKU tels quels :
        #   1) tous remplis
        #   2) tous distincts entre eux
        skus_are_unique = (
            len(non_empty_skus) == len(raw_v_skus)
            and len(set(non_empty_skus)) == len(non_empty_skus)
        )
        if not skus_are_unique and var_count > 1:
            logger.debug(
                f"#{p.id} '{p.name}': SKUs variantes non uniques "
                f"({len(set(non_empty_skus))} distincts sur {var_count}) — génération"
            )

        for idx, vd in enumerate(p.variations_data):
            row = AnkorRow()
            # SKU : règle = si tous uniques, on garde tel quel ; sinon on génère
            v_sku = (vd.get("sku") or "").strip()
            if skus_are_unique and v_sku:
                row.set("SKU", v_sku)
            else:
                row.set("SKU", make_sku(p.slug, p.sku, vd.get("id"), idx))
            # Nom + description (identiques entre variants)
            row.set("Nom du produit", name)
            row.set("Description du produit", cleaned_desc)
            # Attributs : depuis _parent_attributes (déjà mergé dans le scraper)
            attrs = vd.get("_parent_attributes") or []
            sizes, colors, others = [], [], []
            for a in attrs:
                a_name = a.get("name") or ""
                a_value = a.get("value") or ""
                if not a_value:
                    continue
                kind = classify_attribute(a_name)
                if kind == "size":
                    sizes.append(a_value)
                elif kind == "color":
                    colors.append(a_value)
                else:
                    # Pour "autres", on préfixe avec le nom pour ne pas perdre le contexte
                    others.append(f"{a_name}: {a_value}" if a_name else a_value)
            row.set("Tailles des variantes", ", ".join(sizes))
            row.set("Couleurs des variants", ", ".join(colors))
            row.set("Autres attributs de variante", ", ".join(others))
            # Image variation : on ne remplit cette colonne QUE si la variation
            # a réellement sa propre image (différente de Image 1 du parent).
            # Si c'est juste une réplique de l'image parent, on laisse vide.
            v_imgs = extract_image_urls(vd.get("images") or [], limit=1)
            v_image_url = v_imgs[0] if v_imgs else ""
            if v_image_url and parent_images and v_image_url == parent_images[0]:
                v_image_url = ""
            row.set("Image de la variante", v_image_url)
            # Images parent (1..5)
            for i in range(MAX_IMAGES_PER_PRODUCT):
                row.set(f"Image {i+1}", parent_images[i] if i < len(parent_images) else "")
            # Prix
            v_prices = vd.get("prices") or {}
            minor = int(v_prices.get("currency_minor_unit") or 2)
            retail = cents_to_decimal(v_prices.get("price") or v_prices.get("regular_price"), minor)
            row.set("Prix de gros/unité", "")  # rarement exposé par les marques
            row.set("Prix de détail/unité", retail if retail is not None else "")
            row.set("Taux de TVA %", "")
            row.set("Remise sur le prix de gros %", "")
            row.set("Nombre d'unités par paquet", 1)
            # Stock
            row.set("Stock", stock_value(
                bool(vd.get("is_in_stock", True)),
                vd.get("low_stock_remaining"),
            ))
            # Pays + features routées
            row.set("Fabriqué en", feature_cols.get("Fabriqué en", ""))
            row.set("Dimensions", feature_cols.get("Dimensions", ""))
            row.set("Composition", feature_cols.get("Composition", ""))
            row.set("Matériau", feature_cols.get("Matériau", ""))
            row.set("Poids", feature_cols.get("Poids", ""))
            row.set("Volume", feature_cols.get("Volume", ""))
            row.set("Liste INCI", inci_content)
            # Booléens (BOOLEAN_TAGS = 4-tuple : col, anywhere, name_cat_only, negatives)
            for tag_def in BOOLEAN_TAGS:
                col_name = tag_def[0]
                row.set(col_name, "X" if booleans.get(col_name) else "")
            rows.append(row)
    else:
        # Produit simple (ou variable sans variations exploitables)
        row = AnkorRow()
        s_sku = (p.sku or "").strip()
        row.set("SKU", s_sku if s_sku else make_sku(p.slug, p.sku, None, 0))
        row.set("Nom du produit", name)
        row.set("Description du produit", cleaned_desc)
        row.set("Tailles des variantes", "")
        row.set("Couleurs des variants", "")
        row.set("Autres attributs de variante", "")
        row.set("Image de la variante", "")
        for i in range(MAX_IMAGES_PER_PRODUCT):
            row.set(f"Image {i+1}", parent_images[i] if i < len(parent_images) else "")
        minor = int((p.prices or {}).get("currency_minor_unit") or 2)
        retail = cents_to_decimal((p.prices or {}).get("price"), minor)
        row.set("Prix de gros/unité", "")
        row.set("Prix de détail/unité", retail if retail is not None else "")
        row.set("Taux de TVA %", "")
        row.set("Remise sur le prix de gros %", "")
        row.set("Nombre d'unités par paquet", 1)
        row.set("Stock", stock_value(p.is_in_stock, p.low_stock_remaining))
        row.set("Fabriqué en", feature_cols.get("Fabriqué en", ""))
        row.set("Dimensions", feature_cols.get("Dimensions", ""))
        row.set("Composition", feature_cols.get("Composition", ""))
        row.set("Matériau", feature_cols.get("Matériau", ""))
        row.set("Poids", feature_cols.get("Poids", ""))
        row.set("Volume", feature_cols.get("Volume", ""))
        row.set("Liste INCI", inci_content)
        for tag_def in BOOLEAN_TAGS:
            col_name = tag_def[0]
            row.set(col_name, "X" if booleans.get(col_name) else "")
        rows.append(row)

    return rows


# =============================================================================
# SECTION 13 — Écriture du .xlsx Ankorstore
# =============================================================================

def find_template() -> Path | None:
    """Cherche le template Ankorstore dans le dossier du script."""
    here = Path(__file__).parent
    for name in ("ankorstore_template.xlsx", "products (13).xlsx", "products.xlsx"):
        p = here / name
        if p.exists():
            return p
    return None


def write_xlsx(rows: list[AnkorRow], output_path: Path, logger: logging.Logger) -> None:
    """Écrit le xlsx en partant du template Ankorstore si présent, sinon en créant
    un nouveau classeur avec les bons headers.

    Stratégie : on charge le template EN MÉMOIRE (jamais on copy le fichier),
    puis on save vers output_path. Si la destination existe et est locked
    (Excel ouvert dessus), on essaie de la supprimer ; si ça échoue on bascule
    sur un nom alterné avec timestamp pour ne jamais bloquer la marque.
    """
    template = find_template()
    if template:
        logger.info(f"Utilisation du template : {template.name}")
        try:
            wb = openpyxl.load_workbook(template)
        except Exception as e:
            logger.error(f"Impossible de charger le template ({e}) — fallback sur xlsx neuf")
            template = None

    if not template:
        logger.warning("Création d'un xlsx neuf (pas de template ou template illisible)")
        wb = openpyxl.Workbook()
        ws_new = wb.active
        ws_new.title = "Vos produits"
        for idx, col_name in enumerate(ANKORSTORE_COLUMNS, start=1):
            ws_new.cell(row=1, column=idx, value=col_name)

    # Choisit la sheet cible
    if "Vos produits" in wb.sheetnames:
        ws = wb["Vos produits"]
    else:
        ws = wb.active
        logger.warning(f"Sheet 'Vos produits' introuvable, utilisation de '{ws.title}'")
    start_row = 2  # ligne 1 = headers

    # Drop les onglets annexes (LISEZ-MOI, Exemples, Codes pays) : on ne garde
    # que 'Vos produits' dans le fichier final. L'utilisateur a son template à
    # part pour s'y référer ; on évite d'alourdir chaque output.
    for sheet_name in list(wb.sheetnames):
        if sheet_name != ws.title:
            del wb[sheet_name]
            logger.debug(f"Onglet '{sheet_name}' supprimé du fichier de sortie")

    # Mapping nom colonne -> index colonne. Le template a des headers riches type :
    #   "SKU*\nChaque variante doit avoir un SKU unique"
    #   "Fabriqué en (code pays, par ex. FR)\nNécessaire si..."
    # On normalise : 1ère ligne, sans astérisque, sans parenthèses de fin.
    def _normalize_header(s: str) -> str:
        first = str(s).splitlines()[0]
        first = first.rstrip("*").strip()
        # Retire le suffixe entre parenthèses (ex: "Fabriqué en (code pays, par ex. FR)")
        first = re.sub(r"\s*\([^)]+\)\s*$", "", first).strip()
        return first

    def _build_col_to_idx() -> dict[str, int]:
        m: dict[str, int] = {}
        for idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=idx).value or ""
            norm = _normalize_header(header_val)
            m[norm] = idx
            # On indexe aussi par 1re ligne brute (sans dé-parenthèsage) pour permissivité
            first_raw = str(header_val).splitlines()[0].rstrip("*").strip()
            m.setdefault(first_raw, idx)
        return m

    col_to_idx = _build_col_to_idx()

    # Vérifie que toutes nos colonnes cibles sont mappées
    missing = [c for c in ANKORSTORE_COLUMNS if c not in col_to_idx]
    if missing:
        logger.warning(f"Colonnes Ankorstore non trouvées dans le template : {missing}")

    # Expansion auto des colonnes images : si un des produits a > 5 images,
    # on insère des colonnes Image 6, Image 7... jusqu'à 15 maximum.
    max_imgs_used = 5
    for row in rows:
        for i in range(6, MAX_IMAGES_PER_PRODUCT + 1):
            if row.data.get(f"Image {i}"):
                max_imgs_used = max(max_imgs_used, i)
    if max_imgs_used > 5:
        img5_idx = col_to_idx.get("Image 5")
        if img5_idx:
            n_new_cols = max_imgs_used - 5
            ws.insert_cols(img5_idx + 1, amount=n_new_cols)
            # Header pour les nouvelles colonnes (même format que le template)
            IMG_HEADER_FMT = (
                "Image {n}\n"
                "URL ou nom de l'image de la bibliothèque (image.png)\n"
                "Taille : min 500*500px | max 6000*6000px"
            )
            for j in range(n_new_cols):
                new_col_idx = img5_idx + 1 + j
                ws.cell(row=1, column=new_col_idx,
                        value=IMG_HEADER_FMT.format(n=6 + j))
            # Re-calcule col_to_idx puisque toutes les colonnes après Image 5 ont shifté
            col_to_idx = _build_col_to_idx()
            logger.info(
                f"Ajout de {n_new_cols} colonnes images (Image 6 → Image {max_imgs_used})"
            )

    # Remplit les lignes
    for r_idx, row in enumerate(rows, start=start_row):
        for col_name, value in row.data.items():
            idx = col_to_idx.get(col_name)
            if idx is None:
                continue
            ws.cell(row=r_idx, column=idx, value=value)

    # Surlignage jaune des cellules vides/incomplètes sur les champs obligatoires.
    # Permet à la marque (ou au catman) de voir d'un coup d'œil ce qui doit être
    # complété manuellement avant l'import Ankorstore.
    YELLOW_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

    def _is_missing(value, field_name) -> bool:
        """True si la valeur est vide/insuffisante pour ce champ obligatoire."""
        # Stock = 0 est valide (= out of stock)
        if field_name == "Stock":
            return value is None or value == ""
        if value is None or value == "":
            return True
        # Description : min 30 caractères
        if field_name == "Description du produit":
            return len(str(value).strip()) < 30
        return False

    REQUIRED_FIELDS = (
        "SKU", "Nom du produit", "Description du produit", "Image 1",
        "Prix de gros/unité", "Prix de détail/unité", "Taux de TVA %",
        "Nombre d'unités par paquet", "Stock",
    )

    n_highlighted = 0
    end_row = start_row + len(rows) - 1
    for field_name in REQUIRED_FIELDS:
        col_idx = col_to_idx.get(field_name)
        if col_idx is None:
            continue
        for r_idx in range(start_row, end_row + 1):
            value = ws.cell(row=r_idx, column=col_idx).value
            if _is_missing(value, field_name):
                ws.cell(row=r_idx, column=col_idx).fill = YELLOW_FILL
                n_highlighted += 1
    if n_highlighted:
        logger.info(
            f"{n_highlighted} cellule(s) surlignée(s) en jaune "
            f"(champs obligatoires vides ou incomplets à compléter)"
        )

    # Save robuste : si le fichier cible existe et qu'on ne peut pas l'écraser,
    # bascule sur un nom timestampé plutôt que de planter.
    final_path = _safe_save(wb, output_path, logger)
    logger.info(f"Écrit : {final_path} ({len(rows)} lignes / {len(rows)} variantes)")


def _safe_save(wb, output_path: Path, logger: logging.Logger) -> Path:
    """Sauvegarde un workbook openpyxl avec gestion des locks (Excel ouvert)."""
    # 1) Si la cible existe, on essaie de la supprimer avant de save (au cas où
    #    macOS aurait des soucis pour overwrite un fichier locked)
    if output_path.exists():
        try:
            output_path.unlink()
            logger.debug(f"Fichier précédent supprimé : {output_path}")
        except PermissionError:
            logger.warning(
                f"Impossible de supprimer le fichier précédent ({output_path}). "
                f"Probablement ouvert dans Excel / Numbers / Preview — ferme-le et relance, "
                f"ou ne t'inquiète pas : je vais sauver sous un nom alterné."
            )
    # 2) Tentative de save normale
    try:
        wb.save(output_path)
        return output_path
    except PermissionError as e:
        # Bascule sur un nom timestampé
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = output_path.with_name(f"{output_path.stem}_{ts}.xlsx")
        logger.warning(
            f"PermissionError sur {output_path.name} ({e}). "
            f"Sauvegarde sous : {alt.name}"
        )
        wb.save(alt)
        return alt


# =============================================================================
# SECTION 14 — Orchestrateur 1 marque
# =============================================================================

def detect_cms(brand_url: str, logger: logging.Logger) -> str:
    """Détecte le CMS d'un site e-commerce en regardant la page cible (ou la home
    en fallback si l'URL passée est 404).

    Retourne 'squarespace', 'wix', 'prestashop', 'woocommerce' ou 'unknown'.
    """
    body = None
    status = 0
    try:
        status, hdrs, body = http_get_json(brand_url, logger)
    except Exception as e:
        logger.debug(f"Fetch direct échoué : {e}")

    # Si l'URL passée renvoie 404/erreur, on tente la home (racine du domaine)
    # pour quand même détecter le CMS. Utile quand l'utilisateur passe une URL
    # de catégorie qui a changé mais le site fonctionne.
    if status != 200 or not isinstance(body, str):
        parsed = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
        root = f"{parsed.scheme}://{parsed.netloc}/"
        if root != brand_url.rstrip("/") + "/":
            logger.info(
                f"URL passée a renvoyé status={status} → on tente la home {root} "
                f"pour détecter le CMS"
            )
            try:
                status, hdrs, body = http_get_json(root, logger)
            except Exception as e:
                logger.warning(f"Détection CMS impossible (home aussi inaccessible) : {e}")
                return "unknown"

    if status != 200 or not isinstance(body, str):
        logger.warning(f"Détection CMS impossible (status={status})")
        return "unknown"

    body_lower = body.lower()

    # Signatures SumUp Store (très spécifiques)
    sumup_signals = []
    if "sumupstore.com" in body_lower:
        sumup_signals.append("sumupstore.com")
    if "data-selector=\"os-theme-" in body or "data-selector='os-theme-" in body:
        sumup_signals.append("os-theme- selectors")
    if "shop.sumup" in body_lower:
        sumup_signals.append("shop.sumup")
    if "images.sumup.com" in body_lower:
        sumup_signals.append("images.sumup.com")
    if sumup_signals:
        logger.info(f"CMS détecté : SumUp Store ({', '.join(sumup_signals)})")
        return "sumup"

    # Signatures Squarespace (très spécifiques)
    sqsp_signals = []
    if "Static.SQUARESPACE_CONTEXT" in body or "SQUARESPACE_CONTEXT" in body:
        sqsp_signals.append("SQUARESPACE_CONTEXT")
    if "squarespace-cdn.com" in body_lower or "images.squarespace-cdn" in body_lower:
        sqsp_signals.append("squarespace-cdn")
    if "static.squarespace.com" in body_lower:
        sqsp_signals.append("static.squarespace.com")
    if sqsp_signals:
        logger.info(f"CMS détecté : Squarespace ({', '.join(sqsp_signals)})")
        return "squarespace"

    # Signatures Wix (le plus spécifique, on check en premier)
    wix_signals = []
    if "parastorage.com" in body_lower:
        wix_signals.append("parastorage.com (CDN Wix)")
    m_gen = re.search(
        r'<meta[^>]+name=["\']generator["\'][^>]+content=["\']([^"\']*)["\']',
        body, flags=re.IGNORECASE,
    )
    if m_gen and "wix" in m_gen.group(1).lower():
        wix_signals.append(f"meta generator: {m_gen.group(1)}")
    if "_api/wix-" in body or "wix-ecommerce-storefront" in body_lower:
        wix_signals.append("wix-ecommerce-storefront")
    if wix_signals:
        logger.info(f"CMS détecté : Wix ({', '.join(wix_signals)})")
        return "wix"

    # Signatures PrestaShop
    presta_signals = []
    if re.search(r'\bvar\s+prestashop\s*=', body):
        presta_signals.append("var prestashop=")
    if re.search(r'window\.prestashop\b', body):
        presta_signals.append("window.prestashop")
    if "data-prestashop-version" in body_lower:
        presta_signals.append("data-prestashop-version")
    if "prestashop" in body_lower and "/themes/" in body_lower:
        presta_signals.append("themes/ + prestashop string")
    # Meta generator
    m = re.search(r'<meta[^>]+name=["\']generator["\'][^>]+content=["\']([^"\']*)["\']',
                  body, flags=re.IGNORECASE)
    if m and "prestashop" in m.group(1).lower():
        presta_signals.append(f"meta generator: {m.group(1)}")

    # Signatures WooCommerce / WordPress
    woo_signals = []
    if "wp-content" in body_lower:
        woo_signals.append("wp-content/")
    if "/wp-json/" in body_lower or "wp-json\\/" in body_lower:
        woo_signals.append("/wp-json/")
    if re.search(r'\bwoocommerce\b', body, re.IGNORECASE):
        woo_signals.append("woocommerce keyword")
    if "wp-includes" in body_lower:
        woo_signals.append("wp-includes/")
    if m and "wordpress" in m.group(1).lower():
        woo_signals.append(f"meta generator: {m.group(1)}")

    logger.debug(
        f"Signatures CMS détectées — Presta: {presta_signals} | Woo: {woo_signals}"
    )

    # PrestaShop est très spécifique : 1 signature forte suffit
    if presta_signals and len(presta_signals) >= 1:
        # Mais s'il y a aussi des signaux Woo forts (>= 2), on est dubitatif
        # (un site rare peut avoir les deux pour une raison ou une autre)
        if len(woo_signals) >= 3:
            logger.warning(
                f"Signatures mixtes détectées (Presta + Woo). On part sur PrestaShop. "
                f"Presta: {presta_signals} | Woo: {woo_signals}"
            )
        logger.info(f"CMS détecté : PrestaShop ({', '.join(presta_signals)})")
        return "prestashop"

    # WooCommerce : au moins 2 signaux pour être sûr (sinon on pourrait
    # confondre avec un simple site WordPress sans WooCommerce)
    if len(woo_signals) >= 2:
        logger.info(f"CMS détecté : WooCommerce ({', '.join(woo_signals)})")
        return "woocommerce"

    logger.warning(
        f"CMS non détecté. Signaux Presta={presta_signals}, Woo={woo_signals}. "
        f"Utilise --cms woocommerce ou --cms prestashop explicitement."
    )
    return "unknown"


def slugify_domain(brand_url: str) -> str:
    p = urlparse(brand_url if "://" in brand_url else f"https://{brand_url}")
    s = p.netloc.replace("www.", "")
    return re.sub(r"[^a-zA-Z0-9._-]", "_", s)


@dataclass
class BrandReport:
    brand_url: str
    domain: str
    started_at: str
    duration_s: float
    n_products_total: int
    n_products_filtered: int
    n_variants_total: int
    n_warnings: int
    output_file: str
    status: str  # "success" | "partial" | "failed"
    error: str = ""
    # Liste des produits exclus par les filtres : (nom_produit, raison)
    filtered_out_items: list[tuple[str, str]] = field(default_factory=list)


def process_brand(
    brand_url: str,
    output_dir: Path,
    max_products: int | None = None,
    cms: str = "auto",
    concurrency: int = 2,
) -> BrandReport:
    domain_slug = slugify_domain(brand_url)
    brand_dir = output_dir / domain_slug
    brand_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logger(domain_slug, brand_dir)
    started = datetime.now()
    t0 = time.time()
    logger.info(f"=== {brand_url} (CMS: {cms}) ===")

    try:
        # Auto-détection du CMS si demandée
        if cms == "auto":
            detected = detect_cms(brand_url, logger)
            if detected == "unknown":
                raise ValueError(
                    f"CMS non détecté pour {brand_url}. "
                    f"Précise manuellement avec --cms woocommerce ou --cms prestashop."
                )
            cms = detected

        # 1. Scrape — dispatch sur le bon scraper selon le CMS
        if cms == "prestashop":
            scraper = PrestaShopScraper(brand_url, logger)
            products = scraper.build(max_products=max_products, concurrency=concurrency)
        elif cms == "woocommerce":
            scraper = WooScraper(brand_url, logger)
            products = scraper.build(max_products=max_products)
        elif cms == "wix":
            scraper = WixScraper(brand_url, logger)
            products = scraper.build(max_products=max_products, concurrency=concurrency)
        elif cms == "squarespace":
            scraper = SquarespaceScraper(brand_url, logger)
            products = scraper.build(max_products=max_products, concurrency=concurrency)
        elif cms == "custom":
            scraper = CustomScraper(brand_url, logger)
            products = scraper.build(max_products=max_products, concurrency=concurrency)
        elif cms == "sumup":
            scraper = SumUpScraper(brand_url, logger)
            products = scraper.build(max_products=max_products, concurrency=concurrency)
        else:
            raise ValueError(
                f"CMS non supporté : {cms} "
                f"(attendu: woocommerce, prestashop, wix, squarespace, sumup, custom, auto)"
            )
        logger.info(f"Récupéré {len(products)} produits (avant filtre)")

        # 1bis. Normalisation des noms : strip d'un éventuel suffixe site commun
        _strip_common_brand_suffix(products, logger)

        # 1ter. Dédoublonnage par nom normalisé (cas marques qui ont créé 2 fois
        # le même produit, ex: "Monoi passion" et "Monoi Passion")
        products = _dedupe_products_by_name(products, logger)

        # 2. Filtre non-marchandises
        kept: list[WooProduct] = []
        filtered_out: list[tuple[WooProduct, str]] = []
        for p in products:
            is_filtered, reason = is_non_merchandise(p)
            if is_filtered:
                filtered_out.append((p, reason))
                logger.debug(f"FILTRE OUT #{p.id} '{p.name}' ({reason})")
            else:
                kept.append(p)
        logger.info(f"Filtre non-marchandises : {len(filtered_out)} exclus, "
                    f"{len(kept)} retenus")

        # 3. Mapping
        all_rows: list[AnkorRow] = []
        warning_count = 0
        prev_handler_count = len(logger.handlers)
        # On compte les warnings via un compteur custom
        class WarnCounter(logging.Handler):
            def __init__(self): super().__init__(); self.n = 0
            def emit(self, record):
                if record.levelno >= logging.WARNING:
                    self.n += 1
        wc = WarnCounter()
        logger.addHandler(wc)
        for p in kept:
            try:
                rows = build_rows_for_product(p, logger)
                all_rows.extend(rows)
            except Exception as e:
                logger.error(f"#{p.id} '{p.name}' : mapping échoué : {e}")
        warning_count = wc.n

        # 4. Écriture xlsx
        output_file = brand_dir / f"{domain_slug}_ankorstore.xlsx"
        write_xlsx(all_rows, output_file, logger)

        duration = time.time() - t0
        logger.info(f"Terminé en {duration:.1f}s")

        return BrandReport(
            brand_url=brand_url,
            domain=domain_slug,
            started_at=started.isoformat(timespec="seconds"),
            duration_s=round(duration, 1),
            n_products_total=len(products),
            n_products_filtered=len(filtered_out),
            n_variants_total=len(all_rows),
            n_warnings=warning_count,
            output_file=str(output_file),
            status="success" if warning_count == 0 else "partial",
            filtered_out_items=[(p.name, reason) for p, reason in filtered_out],
        )
    except Exception as e:
        logger.exception(f"ÉCHEC pour {brand_url}")
        return BrandReport(
            brand_url=brand_url,
            domain=domain_slug,
            started_at=started.isoformat(timespec="seconds"),
            duration_s=round(time.time() - t0, 1),
            n_products_total=0,
            n_products_filtered=0,
            n_variants_total=0,
            n_warnings=0,
            output_file="",
            status="failed",
            error=str(e),
        )


# =============================================================================
# SECTION 15 — CLI
# =============================================================================

def write_batch_report(reports: list[BrandReport], path: Path) -> None:
    """Écrit un CSV récapitulatif de toutes les marques traitées."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "domain", "brand_url", "status", "n_products_total",
            "n_products_filtered", "n_variants_total", "n_warnings",
            "duration_s", "output_file", "started_at", "error",
        ])
        w.writeheader()
        for r in reports:
            w.writerow({
                "domain": r.domain,
                "brand_url": r.brand_url,
                "status": r.status,
                "n_products_total": r.n_products_total,
                "n_products_filtered": r.n_products_filtered,
                "n_variants_total": r.n_variants_total,
                "n_warnings": r.n_warnings,
                "duration_s": r.duration_s,
                "output_file": r.output_file,
                "started_at": r.started_at,
                "error": r.error,
            })


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Scrape WooCommerce / PrestaShop -> matrice Ankorstore .xlsx",
    )
    parser.add_argument("urls", nargs="+", help="URLs des marques à scraper")
    parser.add_argument("--cms",
                        choices=["auto", "woocommerce", "prestashop", "wix",
                                 "squarespace", "sumup", "custom"],
                        default="auto",
                        help="CMS source (défaut: auto). Force avec un nom explicite "
                             "si la détection se trompe. Use 'custom' pour sites "
                             "sans CMS connu (best-effort via JSON-LD/OG/microdata).")
    parser.add_argument("--max", type=int, default=None,
                        help="Limite le nb de produits par marque (debug)")
    parser.add_argument("--concurrency", type=int, default=2,
                        help="Nb de threads parallèles pour le fetching Presta "
                             "(défaut: 2). Use 1 pour serveurs très lents, "
                             "4-8 pour serveurs rapides.")
    parser.add_argument("--output-dir", default="outputs",
                        help="Dossier de sortie (défaut: ./outputs)")
    args = parser.parse_args()

    here = Path(__file__).parent
    out_dir = (here / args.output_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n>>> Outputs dans : {out_dir} (CMS: {args.cms})\n")

    reports: list[BrandReport] = []
    for url in args.urls:
        report = process_brand(url, out_dir, max_products=args.max,
                               cms=args.cms, concurrency=args.concurrency)
        reports.append(report)
        print()

    # Récap CSV global
    batch_csv = out_dir / "batch_report.csv"
    write_batch_report(reports, batch_csv)

    print("\n" + "=" * 78)
    print(" RÉCAPITULATIF")
    print("=" * 78)
    for r in reports:
        emoji = {"success": "[OK]", "partial": "[WARN]", "failed": "[FAIL]"}.get(r.status, "")
        print(f"  {emoji}  {r.domain}  "
              f"-> {r.n_variants_total} variantes "
              f"(produits {r.n_products_total}/{r.n_products_filtered} filtrés) "
              f"warns={r.n_warnings}  {r.duration_s}s")
        if r.error:
            print(f"        ERROR: {r.error}")
    print(f"\nRapport CSV global : {batch_csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
